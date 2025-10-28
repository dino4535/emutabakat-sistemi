"""
Toplu Mutabakat İşlemleri - Dino Gıda
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Request
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
from backend.database import get_db
from backend.models import User, Mutabakat, MutabakatItem, MutabakatDurumu, UserRole
from backend.schemas import MutabakatResponse
from backend.auth import get_current_active_user
from backend.permissions import Permissions
from backend.logger import ActivityLogger
from backend.middleware.rate_limiter import RateLimiter, RateLimitRules
from pydantic import BaseModel
import random
import string
import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import bcrypt
import re

router = APIRouter(prefix="/api/bulk-mutabakat", tags=["Toplu Mutabakat"])

def generate_mutabakat_no() -> str:
    """Benzersiz mutabakat numarası oluştur"""
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"MUT-{timestamp}-{random_str}"

@router.post("/create-multiple", response_model=List[MutabakatResponse])
def create_bulk_mutabakat(
    request: Request,
    receivers: List[int],  # Müşteri ID listesi
    donem_baslangic: str,
    donem_bitis: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Toplu mutabakat oluştur
    Sadece Admin, Muhasebe ve Planlama kullanabilir
    """
    
    # Yetki kontrolü
    if not Permissions.can_bulk_create(current_user.role):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Toplu mutabakat oluşturma yetkiniz yok"
        )
    
    created_mutabakats = []
    
    for receiver_id in receivers:
        # Alıcı kontrolü
        receiver = db.query(User).filter(User.id == receiver_id).first()
        if not receiver or receiver.role != UserRole.MUSTERI:
            continue
        
        # Mutabakat oluştur (Multi-Company: alıcının company_id'sini kullan)
        mutabakat = Mutabakat(
            company_id=receiver.company_id,  # ÖNEMLİ: Alıcının şirketi
            mutabakat_no=generate_mutabakat_no(),
            sender_id=current_user.id,
            receiver_id=receiver_id,
            receiver_vkn=receiver.vkn_tckn,  # VKN'yi de kaydet
            donem_baslangic=datetime.fromisoformat(donem_baslangic),
            donem_bitis=datetime.fromisoformat(donem_bitis),
            durum=MutabakatDurumu.TASLAK,
            toplam_borc=0.0,
            toplam_alacak=0.0,
            bakiye=0.0
        )
        
        db.add(mutabakat)
        db.commit()
        db.refresh(mutabakat)
        
        created_mutabakats.append(mutabakat)
        
        # Log
        ActivityLogger.log_mutabakat_created(
            db,
            current_user.id,
            mutabakat.mutabakat_no,
            request.client.host if request.client else "unknown"
        )
    
    return created_mutabakats

@router.get("/customers", response_model=List[dict])
def get_customers(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Müşteri listesini getir - Şirket bazlı (Multi-Company)
    Her şirket sadece kendi müşterilerini görebilir
    """
    
    if not Permissions.can_bulk_create(current_user.role):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bu işlem için yetkiniz yok"
        )
    
    # Müşteri sorgusunu hazırla
    customer_query = db.query(User).filter(
        User.role == UserRole.MUSTERI,
        User.is_active == True
    )
    
    # Sistem admini değilse, sadece kendi şirketinin müşterilerini görebilir
    if current_user.role != UserRole.ADMIN:
        customer_query = customer_query.filter(User.company_id == current_user.company_id)
    
    customers = customer_query.all()
    
    return [
        {
            "id": c.id,
            "username": c.username,
            "full_name": c.full_name,
            "company_name": c.company_name,
            "email": c.email
        }
        for c in customers
    ]

# TODO: Netsis entegrasyonu için endpoint
@router.post("/import-from-netsis")
def import_from_netsis(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Netsis'ten cari ekstre çek ve mutabakat oluştur
    (Gelecek özellik)
    """
    
    if not Permissions.can_bulk_create(current_user.role):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bu işlem için yetkiniz yok"
        )
    
    # TODO: Netsis API entegrasyonu
    return {
        "message": "Netsis entegrasyonu yakında eklenecek",
        "status": "coming_soon"
    }


class ExcelUploadResult(BaseModel):
    """Excel yükleme sonucu"""
    toplam: int
    basarili: int
    basarisiz: int
    hatalar: List[dict]
    olusturulan_mutabakatlar: List[dict]


@router.get("/download-template")
def download_excel_template(
    current_user: User = Depends(get_current_active_user)
):
    """
    Excel template dosyasını dinamik olarak oluştur ve indir (Multi-Company)
    """
    if not Permissions.can_bulk_create(current_user.role):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bu işlem için yetkiniz yok"
        )
    
    # Dinamik template oluştur
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Mutabakat Listesi"
    
    # Stil tanımlamaları
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = [
        "Vergi Numarası (VKN/TC)",
        "Bayi Kodu",
        "Bakiye (+ Borç, - Alacak)",
        "Dönem Başlangıç (GG.AA.YYYY)",
        "Dönem Bitiş (GG.AA.YYYY)",
        "Açıklama (Opsiyonel)"
    ]
    
    # Başlıkları yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Örnek veri
    example_data = [
        ["1234567890", "M001", "150000.00", "01.10.2025", "25.10.2025", "Ekim dönemi"],  # Pozitif = Borç
        ["1234567890", "M002", "-50000.00", "01.10.2025", "25.10.2025", ""],  # Negatif = Alacak
        ["9876543210", "T100", "75000.00", "01.10.2025", "25.10.2025", "Ağustos-Ekim dönemi"],   # Başka bir VKN
    ]
    
    # Örnek verileri yaz
    normal_font = Font(size=11)
    normal_alignment = Alignment(horizontal="left", vertical="center")
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = normal_font
            cell.border = border
            cell.alignment = normal_alignment
    
    # Kolon genişlikleri
    column_widths = [22, 18, 20, 22, 22, 35]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Satır yüksekliği
    ws.row_dimensions[1].height = 35
    
    # BytesIO kullanarak Excel'i oluştur
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=Toplu_Mutabakat_Sablonu.xlsx"
        }
    )


def parse_turkish_date(date_str: str) -> Optional[datetime]:
    """Türkçe tarih formatını parse et (GG.AA.YYYY veya GG/AA/YYYY)"""
    if not date_str:
        return None
    
    date_str = str(date_str).strip()
    
    # Excel date serial number ise
    if isinstance(date_str, (int, float)) or date_str.replace('.', '').isdigit():
        try:
            from datetime import timedelta
            # Excel'in epoch'u 1899-12-30
            excel_epoch = datetime(1899, 12, 30)
            return excel_epoch + timedelta(days=float(date_str))
        except:
            pass
    
    # Format denemeleri
    formats = ["%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except:
            continue
    
    return None


def clean_phone(phone: str) -> str:
    """Telefon numarasını temizle ve formatla"""
    if not phone:
        return ""
    
    # Sadece rakamları al
    phone = re.sub(r'[^\d]', '', str(phone))
    
    # Türkiye için 0 ile başlamazsa ekle
    if phone and not phone.startswith('0'):
        phone = '0' + phone
    
    return phone[:11]  # Maksimum 11 karakter


@router.post("/upload-excel", response_model=ExcelUploadResult)
@RateLimiter.limit(**RateLimitRules.EXCEL_UPLOAD)
async def upload_excel_mutabakat(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Excel dosyasından toplu mutabakat yükle (Multi-Company - Bayi Bazlı)
    
    Excel formatı (6 sütun):
    - Vergi Numarası (VKN/TC) - ZORUNLU
    - Bayi Kodu - ZORUNLU
    - Bakiye - ZORUNLU (Pozitif = Borç, Negatif = Alacak)
    - Dönem Başlangıç (GG.AA.YYYY) - ZORUNLU
    - Dönem Bitiş (GG.AA.YYYY) - ZORUNLU
    - Açıklama - Opsiyonel
    
    Mantık:
    - Aynı VKN ve aynı dönem için tek bir mutabakat oluşturulur
    - Her bayi için MutabakatBayiDetay kaydı oluşturulur
    - Bayi tablosundaki bakiye ve son_mutabakat_tarihi güncellenir
    - Mutabakat ALICININ (VKN sahibi) şirketine kaydedilir
    
    NOT:
    - Sistem admini TÜM şirketlere mutabakat oluşturabilir
    - Şirket admini SADECE kendi şirketine mutabakat oluşturabilir
    - VKN'ye göre kullanıcı bulunur/oluşturulur (multi-company aware)
    """
    
    # Yetki kontrolü
    if not Permissions.can_bulk_create(current_user.role):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Toplu mutabakat yükleme yetkiniz yok"
        )
    
    # Dosya uzantısı kontrolü
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Lütfen Excel dosyası (.xlsx veya .xls) yükleyiniz"
        )
    
    try:
        # Excel dosyasını oku
        wb = load_workbook(file.file, data_only=True)
        ws = wb.active
        
        # İlk satır başlıklar
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel dosyası boş veya geçersiz format"
            )
        
        # Maksimum satır kontrolü
        MAX_ROWS = 5000
        if len(rows) > MAX_ROWS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Tek seferde maksimum {MAX_ROWS} satır yüklenebilir. Şu an {len(rows)} satır var."
            )
        
        # VKN + Dönem bazında bayileri grupla
        # Key: (vkn, donem_baslangic, donem_bitis, aciklama)
        # Value: [(bayi_kodu, bakiye), ...]
        mutabakat_groups = {}
        hatalar = []
        
        for row_num, row in enumerate(rows, start=2):
            try:
                # Boş satırı atla
                if not any(row):
                    continue
                
                # Verileri al (Yeni format: 6 sütun)
                vergi_no = str(row[0]).strip() if row[0] else None  # VKN/TC
                bayi_kodu = str(row[1]).strip() if row[1] else None  # Bayi Kodu
                bakiye_str = str(row[2]).strip() if row[2] else "0"  # Bakiye
                donem_baslangic_str = str(row[3]).strip() if len(row) > 3 and row[3] else None  # Dönem Başlangıç
                donem_bitis_str = str(row[4]).strip() if len(row) > 4 and row[4] else None  # Dönem Bitiş
                aciklama = str(row[5]).strip() if len(row) > 5 and row[5] else None  # Açıklama
                
                # Zorunlu alan kontrolü
                if not vergi_no:
                    hatalar.append({
                        "satir": row_num,
                        "hata": "VKN/TC zorunludur"
                    })
                    continue
                
                if not bayi_kodu:
                    hatalar.append({
                        "satir": row_num,
                        "hata": f"Bayi Kodu zorunludur (VKN: {vergi_no})"
                    })
                    continue
                
                if not donem_baslangic_str or not donem_bitis_str:
                    hatalar.append({
                        "satir": row_num,
                        "hata": f"Dönem tarihleri zorunludur (VKN: {vergi_no})"
                    })
                    continue
                
                # Bakiyeyi parse et
                try:
                    bakiye = float(bakiye_str)
                except ValueError:
                    hatalar.append({
                        "satir": row_num,
                        "hata": f"Geçersiz bakiye formatı: {bakiye_str} (VKN: {vergi_no})"
                    })
                    continue
                
                # Tarihleri parse et
                donem_baslangic_dt = parse_turkish_date(donem_baslangic_str)
                donem_bitis_dt = parse_turkish_date(donem_bitis_str)
                
                if not donem_baslangic_dt or not donem_bitis_dt:
                    hatalar.append({
                        "satir": row_num,
                        "hata": f"Geçersiz tarih formatı (GG.AA.YYYY kullanın) - VKN: {vergi_no}"
                    })
                    continue
                
                # VKN + Dönem bazında grupla
                # Key: (vkn, donem_baslangic, donem_bitis, aciklama)
                group_key = (vergi_no, donem_baslangic_dt, donem_bitis_dt, aciklama or "")
                
                if group_key not in mutabakat_groups:
                    mutabakat_groups[group_key] = []
                
                mutabakat_groups[group_key].append((bayi_kodu, bakiye))
                
            except Exception as e:
                hatalar.append({
                    "satir": row_num,
                    "hata": f"Beklenmeyen hata: {str(e)}"
                })
        
        # Şimdi VKN + Dönem bazında mutabakatları oluştur
        from backend.models import Bayi, MutabakatBayiDetay
        from pytz import timezone
        turkey_tz = timezone('Europe/Istanbul')
        
        basarili = 0
        basarisiz = 0
        olusturulan_mutabakatlar = []
        
        for group_key, bayiler in mutabakat_groups.items():
            vergi_no, donem_baslangic_dt, donem_bitis_dt, aciklama = group_key
            try:
                # Kullanıcı var mı kontrol et
                receiver = db.query(User).filter(
                    User.vkn_tckn == vergi_no,
                    User.company_id == current_user.company_id
                ).first()
                
                # Yoksa oluştur
                if not receiver:
                    if len(vergi_no) < 6:
                        hatalar.append({
                            "vkn": vergi_no,
                            "hata": f"VKN çok kısa (minimum 6 karakter): {vergi_no}"
                        })
                        basarisiz += len(bayiler)
                        continue
                    
                    default_password = vergi_no[-6:]
                    hashed_pwd = bcrypt.hashpw(default_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                    
                    receiver = User(
                        company_id=current_user.company_id,
                        username=vergi_no,
                        email=f"{vergi_no}@temp.mutabakat.local",
                        hashed_password=hashed_pwd,
                        full_name=f"Firma {vergi_no}",
                        company_name=f"Firma {vergi_no}",
                        vkn_tckn=vergi_no,
                        role=UserRole.MUSTERI,
                        is_active=True,
                        is_verified=False,
                        ilk_giris_tamamlandi=False,
                        created_at=datetime.now(turkey_tz)
                    )
                    db.add(receiver)
                    db.flush()
                
                # Toplam borç ve alacak hesapla
                toplam_borc = 0.0
                toplam_alacak = 0.0
                
                for _, bakiye in bayiler:
                    if bakiye > 0:
                        toplam_borc += bakiye
                    elif bakiye < 0:
                        toplam_alacak += abs(bakiye)
                
                # Mutabakat oluştur
                mutabakat = Mutabakat(
                    company_id=receiver.company_id,
                    mutabakat_no=generate_mutabakat_no(),
                    sender_id=current_user.id,
                    receiver_id=receiver.id,
                    receiver_vkn=receiver.vkn_tckn,
                    donem_baslangic=donem_baslangic_dt,
                    donem_bitis=donem_bitis_dt,
                    durum=MutabakatDurumu.TASLAK,
                    toplam_borc=toplam_borc,
                    toplam_alacak=toplam_alacak,
                    bakiye=toplam_borc - toplam_alacak,
                    toplam_bayi_sayisi=len(bayiler),
                    aciklama=aciklama if aciklama else None,
                    created_at=datetime.now(turkey_tz)
                )
                
                db.add(mutabakat)
                db.flush()
                
                # Her bayi için detay oluştur ve bayi tablosunu güncelle
                for bayi_kodu, bakiye in bayiler:
                    # MutabakatBayiDetay oluştur
                    detay = MutabakatBayiDetay(
                        mutabakat_id=mutabakat.id,
                        bayi_kodu=bayi_kodu,
                        bayi_adi=f"Bayi {bayi_kodu}",  # Bayi adı bilinmiyor, sadece kod var
                        bakiye=bakiye,
                        created_at=datetime.now(turkey_tz)
                    )
                    db.add(detay)
                    
                    # Bayi tablosunu güncelle (bakiye ve son_mutabakat_tarihi)
                    bayi = db.query(Bayi).filter(
                        Bayi.bayi_kodu == bayi_kodu,
                        Bayi.user_id == receiver.id
                    ).first()
                    
                    if bayi:
                        bayi.bakiye = bakiye
                        bayi.son_mutabakat_tarihi = datetime.now(turkey_tz)
                        bayi.updated_at = datetime.now(turkey_tz)
                
                # Log
                ActivityLogger.log_mutabakat_created(
                    db,
                    current_user.id,
                    mutabakat.mutabakat_no,
                    request.client.host if request.client else "unknown"
                )
                
                olusturulan_mutabakatlar.append({
                    "mutabakat_no": mutabakat.mutabakat_no,
                    "vkn": vergi_no,
                    "bayi_sayisi": len(bayiler),
                    "toplam_borc": toplam_borc,
                    "toplam_alacak": toplam_alacak,
                    "bakiye": mutabakat.bakiye,
                    "donem": f"{donem_baslangic_dt.strftime('%d.%m.%Y')} - {donem_bitis_dt.strftime('%d.%m.%Y')}",
                    "aciklama": aciklama if aciklama else "-"
                })
                
                basarili += 1
                
            except Exception as e:
                hatalar.append({
                    "vkn": vergi_no,
                    "hata": f"Mutabakat oluşturma hatası: {str(e)}"
                })
                basarisiz += 1
        
        # Tüm işlemler başarılı ise commit et
        db.commit()
        
        return ExcelUploadResult(
            toplam=len(mutabakat_groups),  # Toplam mutabakat sayısı (VKN + Dönem kombinasyonları)
            basarili=basarili,
            basarisiz=basarisiz,
            hatalar=hatalar,
            olusturulan_mutabakatlar=olusturulan_mutabakatlar
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel işleme hatası: {str(e)}"
        )


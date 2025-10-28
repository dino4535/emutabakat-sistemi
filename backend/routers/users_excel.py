"""
Kullanıcı Excel İşlemleri
Toplu kullanıcı yükleme ve template oluşturma
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Request
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
from backend.database import get_db
from backend.models import User, UserRole
from backend.auth import get_current_active_user
from backend.logger import ActivityLogger
from backend.middleware.rate_limiter import RateLimiter, RateLimitRules
from pydantic import BaseModel
import bcrypt
import random
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

router = APIRouter(prefix="/api/auth", tags=["Kullanıcı Excel"])


class ExcelUserUploadResult(BaseModel):
    toplam: int
    basarili_user: int
    basarili_bayi: int
    basarisiz: int
    hatalar: List[dict]
    olusturulan_kullanicilar: List[dict]
    olusturulan_bayiler: List[dict]


def create_users_template():
    """Kullanıcı Excel template oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kullanıcı Listesi"
    
    # Başlık stili
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = [
        "Vergi Numarası (VKN/TC)",
        "Bayi Kodu",
        "Bayi Adı (Şube/Mağaza)",
        "E-posta (Opsiyonel - İlk kayıt için)",
        "Şifre (Opsiyonel - Boşsa VKN Son 6 Hane)",
        "Telefon (Opsiyonel)",
        "Adres (Opsiyonel)",
        "Rol (musteri/tedarikci)",
        "Şirket (Sadece Sistem Admini İçin - VKN veya Şirket Adı)"
    ]
    
    # Başlıkları yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Örnek veriler (VKN bazlı - Aynı VKN birden fazla bayi kodu ile olabilir)
    # NOT: 
    # - Kullanıcı adı otomatik olarak VKN'den oluşur
    # - Email ve telefon opsiyonel (ilk girişte profilde doldurulur)
    # - Şifre opsiyonel (boşsa VKN son 6 hane kullanılır)
    example_data = [
        [
            "1234567890",           # VKN/TC (ZORUNLU) - Kullanıcı adı olarak da kullanılır
            "M001",                 # Bayi Kodu (ZORUNLU)
            "ABC Gıda Merkez Mağaza",  # Bayi Adı (ZORUNLU)
            "",                     # E-posta (Opsiyonel - boş bırakılabilir)
            "",                     # Şifre (Opsiyonel - Boşsa VKN son 6 hane: 567890)
            "",                     # Telefon (Opsiyonel)
            "İzmir Menderes",       # Adres (Opsiyonel)
            "musteri",              # Rol
            "Dino Gıda"             # Şirket (Opsiyonel)
        ],
        [
            "1234567890",           # Aynı VKN - İkinci bayi kodu
            "M002",                 # Farklı Bayi Kodu
            "ABC Gıda Bornova Şubesi",  # Farklı Bayi Adı
            "",                     # E-posta boş
            "",                     # Şifre boş (kullanıcı zaten var)
            "",                     # Telefon boş
            "İzmir Bornova",        # Adres
            "musteri",              # Rol
            "Dino Gıda"             # Şirket
        ],
        [
            "9876543210",           # Farklı VKN (İlk kayıt)
            "T100",                 # Bayi Kodu
            "XYZ Tedarik Ana Depo", # Bayi Adı
            "",                     # E-posta (Opsiyonel - ilk girişte profilde doldurulur)
            "",                     # Şifre (Opsiyonel - Boşsa VKN son 6 hane: 543210)
            "",                     # Telefon (Opsiyonel)
            "İzmir Karşıyaka",      # Adres
            "tedarikci",            # Rol
            "Bermer"                # Şirket
        ]
    ]
    
    # Örnek verileri yaz
    normal_font = Font(size=11)
    normal_alignment = Alignment(horizontal="left", vertical="center")
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = normal_font
            cell.border = border
            cell.alignment = normal_alignment
    
    # Kolon genişlikleri (VKN bazlı - Kullanıcı Adı kolonu kaldırıldı)
    column_widths = [20, 15, 30, 30, 15, 15, 25, 25, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Satır yüksekliği
    ws.row_dimensions[1].height = 40
    
    # Kaydet
    template_path = Path("template_kullanicilar.xlsx")
    wb.save(template_path)
    return str(template_path)


@router.get("/download-user-template")
def download_user_template(
    current_user: User = Depends(get_current_active_user)
):
    """
    Kullanıcı Excel template indir (Sistem ve Şirket Admini)
    """
    if current_user.role not in [UserRole.ADMIN, UserRole.COMPANY_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bu işlem için yetkiniz yok"
        )
    
    # Template oluştur veya mevcut olanı kullan
    template_path = Path("template_kullanicilar.xlsx")
    if not template_path.exists():
        create_users_template()
    
    return FileResponse(
        path=str(template_path),
        filename="Kullanici_Sablonu.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def clean_phone(phone: str) -> str:
    """Telefon numarasını temizle"""
    if not phone:
        return ""
    phone = re.sub(r'[^\d]', '', str(phone))
    if phone and not phone.startswith('0'):
        phone = '0' + phone
    return phone[:11]


def get_company_slug(company_name: str) -> str:
    """Şirket adından slug oluştur (VKN_CompanySlug formatı için)"""
    # Türkçe karakterleri değiştir
    tr_map = {
        'ı': 'i', 'İ': 'i', 'ş': 's', 'Ş': 's',
        'ğ': 'g', 'Ğ': 'g', 'ü': 'u', 'Ü': 'u',
        'ö': 'o', 'Ö': 'o', 'ç': 'c', 'Ç': 'c'
    }
    
    slug = company_name.lower()
    for tr_char, en_char in tr_map.items():
        slug = slug.replace(tr_char, en_char)
    
    # Sadece alfanumerik karakterleri tut
    slug = re.sub(r'[^a-z0-9]', '', slug)
    
    # İlk 15 karakteri al (çok uzun olmasın)
    return slug[:15] if slug else 'company'


@router.post("/upload-users-excel", response_model=ExcelUserUploadResult)
@RateLimiter.limit(**RateLimitRules.EXCEL_UPLOAD)
async def upload_users_excel(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Excel dosyasından toplu kullanıcı ve bayi yükle - VKN Bazlı Sistem (Sistem ve Şirket Admini)
    
    Excel formatı:
    1. Vergi Numarası (VKN/TC) - ZORUNLU (Kullanıcı adı olarak da kullanılır)
    2. Bayi Kodu - ZORUNLU
    3. Bayi Adı (Şube/Mağaza) - ZORUNLU
    4. E-posta - Opsiyonel (İlk girişte profilde doldurulur)
    5. Şifre - İlk VKN kaydı için ZORUNLU
    6. Telefon - Opsiyonel (İlk girişte profilde doldurulur)
    7. Adres - Opsiyonel
    8. Rol (musteri/tedarikci) - ZORUNLU
    9. Şirket (Sadece Sistem Admini İçin - VKN veya Şirket Adı) - Sistem admini için ZORUNLU
    
    NOT: 
    - Kullanıcı adı OTOMATIK olarak VKN numarası olur
    - E-posta ve telefon ilk girişte KVKK onayından sonra profilde doldurulur
    - Bir VKN'ye ait birden fazla bayi kodu (şube/mağaza) tanımlanabilir
    - Aynı VKN için ilk satırda sadece şifre zorunlu (email/telefon opsiyonel)
    - Aynı VKN için sonraki satırlarda kullanıcı bilgileri boş bırakılabilir (sadece bayi eklenir)
    - Sistem admini için 'Şirket' kolonu ZORUNLUDUR
    - Şirket admini için 'Şirket' kolonu boş bırakılabilir (otomatik kendi şirketi atanır)
    """
    
    # Yetki kontrolü (Multi-Company)
    if current_user.role not in [UserRole.ADMIN, UserRole.COMPANY_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Toplu kullanıcı yükleme yetkiniz yok"
        )
    
    # Dosya uzantısı kontrolü
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Lütfen Excel dosyası (.xlsx veya .xls) yükleyiniz"
        )
    
    try:
        # Excel dosyasını oku
        wb = load_workbook(file.file, data_only=True)
        ws = wb.active
        
        # İlk satır başlıklar
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel dosyası boş veya geçersiz format"
            )
        
        # Maksimum satır kontrolü
        MAX_ROWS = 1000
        if len(rows) > MAX_ROWS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Tek seferde maksimum {MAX_ROWS} satır yüklenebilir. Şu an {len(rows)} satır var."
            )
        
        basarili_user = 0
        basarili_bayi = 0
        basarisiz = 0
        hatalar = []
        olusturulan_kullanicilar = []
        olusturulan_bayiler = []
        
        # VKN Cache: Her VKN için bir kez user oluştur, sonra sadece bayi ekle
        vkn_user_cache = {}
        
        # Rol mapping (sadece musteri/tedarikci)
        rol_mapping = {
            "musteri": UserRole.MUSTERI,
            "müşteri": UserRole.MUSTERI,
            "tedarikci": UserRole.TEDARIKCI,
            "tedarikçi": UserRole.TEDARIKCI
        }
        
        # Import Bayi model
        from backend.models import Bayi
        
        for row_num, row in enumerate(rows, start=2):
            vkn_tckn = None  # Exception handling için
            
            # Her satır için savepoint oluştur (hata durumunda sadece bu satırı geri al)
            savepoint = db.begin_nested()
            
            try:
                # Boş satırı atla
                if not any(row):
                    savepoint.commit()  # Boş satır için savepoint'i kapat
                    continue
                
                # Verileri al (VKN bazlı)
                vkn_tckn = str(row[0]).strip() if row[0] else None  # VKN/TC (ZORUNLU)
                bayi_kodu = str(row[1]).strip() if row[1] else None  # Bayi Kodu (ZORUNLU)
                bayi_adi = str(row[2]).strip() if row[2] else None  # Bayi Adı (ZORUNLU)
                email = str(row[3]).strip().lower() if row[3] else None  # E-posta (İlk kayıt için)
                password = str(row[4]).strip() if row[4] else None  # Şifre (İlk kayıt için)
                phone = clean_phone(row[5]) if row[5] else None  # Telefon
                address = str(row[6]).strip() if row[6] else None  # Adres
                rol_str = str(row[7]).strip().lower() if row[7] else "musteri"  # Rol
                company_identifier = str(row[8]).strip() if len(row) > 8 and row[8] else None  # Şirket
                
                # Otomatik değerler
                tax_number = vkn_tckn  # tax_number = vkn_tckn
                full_name = bayi_adi  # Ad Soyad = Bayi Adı
                company_name = bayi_adi  # Şirket Adı = Bayi Adı
                
                # Zorunlu alan kontrolü
                if not vkn_tckn:
                    savepoint.rollback()
                    hatalar.append({
                        "satir": row_num,
                        "hata": "VKN/TC zorunludur",
                        "vkn": "-"
                    })
                    basarisiz += 1
                    continue
                
                if not bayi_kodu:
                    savepoint.rollback()
                    hatalar.append({
                        "satir": row_num,
                        "hata": "Bayi Kodu zorunludur",
                        "vkn": vkn_tckn
                    })
                    basarisiz += 1
                    continue
                
                if not bayi_adi:
                    savepoint.rollback()
                    hatalar.append({
                        "satir": row_num,
                        "hata": "Bayi Adı zorunludur",
                        "vkn": vkn_tckn
                    })
                    basarisiz += 1
                    continue
                
                # E-posta formatı kontrolü (eğer varsa)
                if email:
                    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                    if not re.match(email_regex, email):
                        savepoint.rollback()
                        hatalar.append({
                            "satir": row_num,
                            "hata": f"Geçersiz e-posta formatı: {email}",
                            "vkn": vkn_tckn
                        })
                        basarisiz += 1
                        continue
                
                # Rol kontrolü
                if rol_str not in rol_mapping:
                    savepoint.rollback()
                    hatalar.append({
                        "satir": row_num,
                        "hata": f"Geçersiz rol: {rol_str}. Olası değerler: musteri, tedarikci",
                        "vkn": vkn_tckn
                    })
                    basarisiz += 1
                    continue
                
                role = rol_mapping[rol_str]
                
                # Şirket ID belirleme (Multi-Company)
                from backend.models import Company
                
                target_company_id = None
                target_company = None
                
                if current_user.role == UserRole.ADMIN:
                    # Sistem admini: Şirket kolonu zorunlu
                    if not company_identifier:
                        savepoint.rollback()
                        hatalar.append({
                            "satir": row_num,
                            "hata": "Sistem admini için 'Şirket' kolonu zorunludur (VKN veya şirket adı giriniz)",
                            "vkn": vkn_tckn
                        })
                        basarisiz += 1
                        continue
                    
                    # Şirket VKN veya adına göre bul
                    target_company = db.query(Company).filter(
                        (Company.vkn == company_identifier) | (Company.company_name.ilike(f"%{company_identifier}%"))
                    ).first()
                    
                    if not target_company:
                        savepoint.rollback()
                        hatalar.append({
                            "satir": row_num,
                            "hata": f"Şirket bulunamadı: {company_identifier}. Lütfen VKN veya şirket adını doğru giriniz.",
                            "vkn": vkn_tckn
                        })
                        basarisiz += 1
                        continue
                    
                    target_company_id = target_company.id
                else:
                    # Şirket admini: Otomatik olarak kendi şirketine atar
                    target_company_id = current_user.company_id
                    # Company bilgisini al
                    target_company = db.query(Company).filter(Company.id == target_company_id).first()
                
                # Username oluştur: VKN_CompanySlug formatında
                company_slug = get_company_slug(target_company.company_name)
                username = f"{vkn_tckn}_{company_slug}"
                
                # VKN + Company ID ile cache key oluştur
                cache_key = f"{vkn_tckn}_{target_company_id}"
                
                # Bu VKN için user var mı kontrol et (Cache veya DB)
                if cache_key not in vkn_user_cache:
                    # Database'de var mı? (VKN + Company ID kontrolü)
                    existing_user = db.query(User).filter(
                        User.vkn_tckn == vkn_tckn,
                        User.company_id == target_company_id
                    ).first()
                    
                    # Username ile de kontrol et (güvenlik için)
                    if not existing_user:
                        existing_user = db.query(User).filter(
                            User.username == username
                        ).first()
                    
                    if existing_user:
                        # Varolan kullanıcıyı cache'le
                        vkn_user_cache[cache_key] = existing_user
                    else:
                        # İlk kayıt: Şifre yoksa VKN'nin son 6 hanesini kullan
                        if not password:
                            if len(vkn_tckn) < 6:
                                savepoint.rollback()
                                hatalar.append({
                                    "satir": row_num,
                                    "hata": f"VKN/TC çok kısa (minimum 6 karakter gerekli): {vkn_tckn}",
                                    "vkn": vkn_tckn
                                })
                                basarisiz += 1
                                continue
                            password = vkn_tckn[-6:]  # Otomatik şifre: VKN son 6 hane
                        
                        # Email varsa unique kontrolü yap (opsiyonel)
                        if email:
                            dup_email = db.query(User).filter(User.email == email).first()
                            if dup_email:
                                savepoint.rollback()
                                hatalar.append({
                                    "satir": row_num,
                                    "hata": f"Bu e-posta zaten kullanılıyor: {email}",
                                    "vkn": vkn_tckn
                                })
                                basarisiz += 1
                                continue
                        
                        # Şifreyi hashle
                        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                        
                        # Kullanıcı oluştur (Multi-Company)
                        # Email ve telefon ilk girişte profilde doldurulacak
                        new_user = User(
                            company_id=target_company_id,
                            username=username,
                            email=email if email else f"{username}@temp.mutabakat.local",  # Geçici email (unique olmalı)
                            hashed_password=hashed_password,
                            full_name=full_name,
                            company_name=company_name,
                            vkn_tckn=tax_number,
                            phone=phone,
                            address=address,
                            role=role,
                            is_active=True,
                            is_verified=False,
                            ilk_giris_tamamlandi=False,  # İlk girişte profil doldurulacak
                            created_at=datetime.utcnow()
                        )
                        
                        db.add(new_user)
                        db.flush()
                        
                        vkn_user_cache[cache_key] = new_user
                        
                        olusturulan_kullanicilar.append({
                            "vkn": vkn_tckn,
                            "username": username,
                            "email": email,
                            "full_name": full_name,
                            "role": role.value
                        })
                        
                        basarili_user += 1
                
                # User referansı al
                user = vkn_user_cache[cache_key]
                
                # Bayi kaydı oluştur
                # Aynı bayi kodu var mı kontrol et
                existing_bayi = db.query(Bayi).filter(
                    Bayi.user_id == user.id,
                    Bayi.bayi_kodu == bayi_kodu
                ).first()
                
                if existing_bayi:
                    savepoint.rollback()
                    hatalar.append({
                        "satir": row_num,
                        "hata": f"Bu bayi kodu zaten kayıtlı: {bayi_kodu} (VKN: {vkn_tckn})",
                        "vkn": vkn_tckn
                    })
                    basarisiz += 1
                    continue
                
                new_bayi = Bayi(
                    user_id=user.id,
                    bayi_kodu=bayi_kodu,
                    bayi_adi=bayi_adi,
                    vkn_tckn=vkn_tckn,
                    bakiye=0.0,
                    donem=None,
                    created_at=datetime.utcnow()
                )
                
                db.add(new_bayi)
                db.flush()
                
                olusturulan_bayiler.append({
                    "vkn": vkn_tckn,
                    "bayi_kodu": bayi_kodu,
                    "bayi_adi": bayi_adi
                })
                
                basarili_bayi += 1
                
                # Savepoint'i commit et (başarılı satır)
                savepoint.commit()
                
            except ValueError as e:
                savepoint.rollback()  # Sadece bu satırı geri al
                hatalar.append({
                    "satir": row_num,
                    "hata": f"Geçersiz veri formatı: {str(e)}",
                    "vkn": vkn_tckn if vkn_tckn else "Bilinmiyor"
                })
                basarisiz += 1
            except Exception as e:
                savepoint.rollback()  # Sadece bu satırı geri al
                hatalar.append({
                    "satir": row_num,
                    "hata": f"Beklenmeyen hata: {str(e)}",
                    "vkn": vkn_tckn if vkn_tckn else "Bilinmiyor"
                })
                basarisiz += 1
        
        # Tüm işlemler başarılı ise commit et
        db.commit()
        
        # Log
        ActivityLogger.log_activity(
            db,
            current_user.id,
            "TOPLU_VKN_BAYI_YUKLE",
            f"{basarili_user} kullanıcı, {basarili_bayi} bayi Excel'den yüklendi (VKN bazlı)",
            request.client.host if request.client else "unknown"
        )
        
        return ExcelUserUploadResult(
            toplam=len(rows),
            basarili_user=basarili_user,
            basarili_bayi=basarili_bayi,
            basarisiz=basarisiz,
            hatalar=hatalar,
            olusturulan_kullanicilar=olusturulan_kullanicilar,
            olusturulan_bayiler=olusturulan_bayiler
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel işleme hatası: {str(e)}"
        )


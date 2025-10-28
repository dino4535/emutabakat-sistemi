"""
Bayi Yönetimi Router - VKN Bazlı Çoklu Bayi Yönetimi
"""

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime
from io import BytesIO
import openpyxl
from pathlib import Path
import re
import bcrypt
import json

from backend.database import get_db
from backend.models import User, Bayi, UserRole
from backend.routers.auth import get_current_active_user
from backend.logger import ActivityLogger

router = APIRouter(
    prefix="/api/bayi",
    tags=["bayi"]
)

# =====================
# SCHEMAS
# =====================

class BayiCreate(BaseModel):
    bayi_kodu: str
    vkn_tckn: str
    bayi_adi: str
    bakiye: float = 0.0
    donem: Optional[str] = None
    adres: Optional[str] = None
    il: Optional[str] = None
    ilce: Optional[str] = None

class BayiResponse(BaseModel):
    id: int
    bayi_kodu: str
    vkn_tckn: str
    bayi_adi: str
    bakiye: float
    donem: Optional[str]
    user_id: int
    created_at: datetime
    
    class Config:
        from_attributes = True

class BayiSummaryByVKN(BaseModel):
    vkn_tckn: str
    company_name: Optional[str]
    toplam_bayi_sayisi: int
    toplam_bakiye: float
    bayiler: List[BayiResponse]

class ExcelUploadResult(BaseModel):
    toplam: int
    basarili: int
    basarisiz: int
    hatalar: List[dict]
    olusturulan_bayiler: List[dict]

# =====================
# HELPER FUNCTIONS
# =====================

def validate_vkn_tckn(vkn_tckn: str) -> bool:
    """VKN (10 hane) veya TC (11 hane) doğrula"""
    if not vkn_tckn:
        return False
    
    # Sadece rakamlardan oluşmalı
    if not vkn_tckn.isdigit():
        return False
    
    # 10 veya 11 haneli olmalı
    if len(vkn_tckn) not in [10, 11]:
        return False
    
    return True

def parse_donem(donem_str: str) -> Optional[str]:
    """Dönem string'ini YYYY-MM formatına çevir"""
    if not donem_str:
        return None
    
    donem_str = str(donem_str).strip()
    
    # YYYY-MM formatı
    if re.match(r'^\d{4}-\d{2}$', donem_str):
        return donem_str
    
    # MM/YYYY formatı
    if re.match(r'^\d{1,2}/\d{4}$', donem_str):
        parts = donem_str.split('/')
        return f"{parts[1]}-{parts[0].zfill(2)}"
    
    # YYYY/MM formatı
    if re.match(r'^\d{4}/\d{1,2}$', donem_str):
        parts = donem_str.split('/')
        return f"{parts[0]}-{parts[1].zfill(2)}"
    
    return donem_str

def get_or_create_user_by_vkn(db: Session, vkn_tckn: str, company_id: int, company_name: str = None, tax_office: str = None) -> User:
    """VKN'ye göre kullanıcı bul veya oluştur (Multi-Company)"""
    
    # Önce VKN ve company_id'ye göre ara (Multi-Company: aynı VKN farklı şirketlerde olabilir)
    user = db.query(User).filter(
        User.vkn_tckn == vkn_tckn,
        User.company_id == company_id
    ).first()
    
    if user:
        # Eğer kullanıcı varsa ve vergi dairesi bilgisi yoksa güncelle
        if tax_office and not user.tax_office:
            user.tax_office = tax_office
            db.commit()
            db.refresh(user)
        return user
    
    # Kullanıcı yok, yeni oluştur
    # Şifre: VKN'nin son 6 hanesi
    password = vkn_tckn[-6:]
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    # Geçici email oluştur (ilk giriş'te değiştirilecek)
    # Email unique olmalı, o yüzden şirket ID'sini de ekleyelim
    temp_email = f"{vkn_tckn}.{company_id}@temp.mutabakat.local"
    
    user = User(
        company_id=company_id,  # ÇOK ÖNEMLİ: Şirket ID'sini set et!
        vkn_tckn=vkn_tckn,
        username=f"{vkn_tckn}_{company_id}",  # Username unique olmalı (aynı VKN farklı şirketlerde olabilir)
        hashed_password=hashed_password,
        email=temp_email,  # Geçici email (ilk giriş'te güncellenecek)
        company_name=company_name or f"Firma {vkn_tckn}",
        tax_office=tax_office,  # Vergi dairesi
        role=UserRole.MUSTERI,
        is_active=True,
        is_verified=False,
        ilk_giris_tamamlandi=False
    )
    
    db.add(user)
    db.commit()
    db.refresh(user)
    
    print(f"[BAYI] Yeni kullanici olusturuldu: VKN={vkn_tckn}, Email={temp_email}, Sifre={password} (son 6 hane)")
    
    return user

# =====================
# ENDPOINTS
# =====================

@router.get("/download-template")
def download_bayi_template(
    current_user: User = Depends(get_current_active_user)
):
    """
    Bayi toplu yükleme için Excel template'i indir (Multi-Company)
    """
    # Sadece admin, company_admin, muhasebe, planlama indirebilir
    if current_user.role not in [UserRole.ADMIN, UserRole.COMPANY_ADMIN, UserRole.MUHASEBE, UserRole.PLANLAMA]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bu işlem için yetkiniz yok"
        )
    
    template_path = Path("template_bayiler.xlsx")
    
    if not template_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Template dosyası bulunamadı. Lütfen 'py create_bayi_template.py' çalıştırın."
        )
    
    return FileResponse(
        path=str(template_path),
        filename="Bayi_Yukleme_Sablonu.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.post("/upload-excel", response_model=ExcelUploadResult)
def upload_excel_bayi(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Excel dosyasından toplu bayi yükle (Multi-Company)
    
    Excel Formatı:
    - Bayi Kodu (zorunlu, unique)
    - VKN / TC Kimlik (zorunlu, 10 veya 11 hane)
    - Bayi Adı (zorunlu)
    - Bakiye (zorunlu, sayı)
    - Dönem (YYYY-MM formatında)
    """
    
    # Yetki kontrolü
    if current_user.role not in [UserRole.ADMIN, UserRole.COMPANY_ADMIN, UserRole.MUHASEBE, UserRole.PLANLAMA]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bu işlem için yetkiniz yok"
        )
    
    # Dosya uzantısı kontrolü
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sadece Excel dosyası (.xlsx, .xls) yüklenebilir"
        )
    
    try:
        # Excel'i oku
        contents = file.file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Başlık satırını atla, veriler 2. satırdan başlar
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel dosyası boş"
            )
        
        # Maksimum satır kontrolü
        MAX_ROWS = 5000
        if len(rows) > MAX_ROWS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Tek seferde maksimum {MAX_ROWS} satır yüklenebilir. Şu an {len(rows)} satır var."
            )
        
        # Sonuçlar
        basarili_count = 0
        basarisiz_count = 0
        hatalar = []
        olusturulan_bayiler = []
        
        # Her satırı işle
        for row_num, row in enumerate(rows, start=2):
            try:
                # Boş satırı atla
                if not any(row):
                    continue
                
                # Kolonları parse et
                bayi_kodu = str(row[0]).strip() if row[0] else None
                vkn_tckn = str(row[1]).strip() if row[1] else None
                bayi_adi = str(row[2]).strip() if row[2] else None
                bakiye_str = row[3]
                donem_str = str(row[4]).strip() if row[4] else None
                
                # Validasyon
                if not bayi_kodu:
                    raise ValueError("Bayi Kodu boş olamaz")
                
                if not vkn_tckn:
                    raise ValueError("VKN / TC Kimlik boş olamaz")
                
                if not validate_vkn_tckn(vkn_tckn):
                    raise ValueError(f"Geçersiz VKN/TC: {vkn_tckn} (10 veya 11 haneli rakam olmalı)")
                
                if not bayi_adi:
                    raise ValueError("Bayi Adı boş olamaz")
                
                # Bakiye parse
                try:
                    bakiye = float(bakiye_str) if bakiye_str is not None else 0.0
                except:
                    raise ValueError(f"Geçersiz bakiye: {bakiye_str}")
                
                # Dönem parse
                donem = parse_donem(donem_str) if donem_str else None
                
                # Bayi kodu unique mi?
                existing_bayi = db.query(Bayi).filter(Bayi.bayi_kodu == bayi_kodu).first()
                if existing_bayi:
                    raise ValueError(f"Bayi kodu zaten mevcut: {bayi_kodu}")
                
                # VKN'ye göre user bul veya oluştur (Multi-Company: current_user'ın company_id'sini kullan)
                user = get_or_create_user_by_vkn(db, vkn_tckn, current_user.company_id, bayi_adi)
                
                # Yeni bayi oluştur
                new_bayi = Bayi(
                    bayi_kodu=bayi_kodu,
                    vkn_tckn=vkn_tckn,
                    user_id=user.id,
                    bayi_adi=bayi_adi,
                    bakiye=bakiye,
                    donem=donem
                )
                
                db.add(new_bayi)
                db.flush()  # ID'yi al ama henüz commit etme
                
                basarili_count += 1
                olusturulan_bayiler.append({
                    "satir": row_num,
                    "bayi_kodu": bayi_kodu,
                    "vkn_tckn": vkn_tckn,
                    "bayi_adi": bayi_adi,
                    "bakiye": bakiye
                })
                
            except Exception as row_error:
                basarisiz_count += 1
                hatalar.append({
                    "satir": row_num,
                    "hata": str(row_error),
                    "veri": {
                        "bayi_kodu": row[0] if len(row) > 0 else None,
                        "vkn_tckn": row[1] if len(row) > 1 else None,
                        "bayi_adi": row[2] if len(row) > 2 else None
                    }
                })
        
        # Tüm işlemler başarılıysa commit et
        if basarili_count > 0:
            db.commit()
            
            # Activity log
            ActivityLogger.log(
                db=db,
                action="TOPLU_BAYI_YUKLE",
                description=f"{basarili_count} adet bayi yüklendi (Excel)",
                user_id=current_user.id,
                ip_address="127.0.0.1"
            )
        else:
            db.rollback()
        
        return ExcelUploadResult(
            toplam=len(rows),
            basarili=basarili_count,
            basarisiz=basarisiz_count,
            hatalar=hatalar,
            olusturulan_bayiler=olusturulan_bayiler
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"[HATA] Excel upload hatasi: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Beklenmeyen hata: {str(e)[:100]}"
        )

@router.post("/upload-excel-stream")
async def upload_excel_bayi_stream(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Excel'den bayi yükle - Gerçek zamanlı ilerleme ile (Streaming, Multi-Company)
    """
    # Yetki kontrolü
    if current_user.role not in [UserRole.ADMIN, UserRole.COMPANY_ADMIN, UserRole.MUHASEBE]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bu işlem için yetkiniz yok"
        )
    
    # Dosya uzantısı kontrolü
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sadece Excel dosyası (.xlsx, .xls) yüklenebilir"
        )
    
    async def generate_progress():
        try:
            # Excel'i oku
            contents = await file.read()
            wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
            ws = wb.active
            
            # Başlık satırını atla, veriler 2. satırdan başlar
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            if not rows:
                yield f"data: {json.dumps({'type': 'error', 'message': 'Excel dosyası boş'})}\n\n"
                return
            
            # Maksimum satır kontrolü
            MAX_ROWS = 5000
            if len(rows) > MAX_ROWS:
                yield f"data: {json.dumps({'type': 'error', 'message': f'Tek seferde maksimum {MAX_ROWS} satır yüklenebilir'})}\n\n"
                return
            
            total_rows = len(rows)
            yield f"data: {json.dumps({'type': 'total', 'count': total_rows})}\n\n"
            
            # Sonuçlar
            basarili_count = 0
            basarisiz_count = 0
            hatalar = []
            olusturulan_bayiler = []
            
            # Batch işleme (10'ar 10'ar commit)
            BATCH_SIZE = 10
            batch_counter = 0
            
            # Her satırı işle
            for row_num, row in enumerate(rows, start=2):
                try:
                    # Boş satırı atla
                    if not any(row):
                        continue
                    
                    # 4 kolon: VKN/TC, Bayi Kodu, Bayi Adı, Vergi Dairesi
                    vkn_tckn = str(row[0]).strip() if row[0] else None
                    bayi_kodu = str(row[1]).strip() if row[1] else None
                    bayi_adi = str(row[2]).strip() if row[2] else None
                    vergi_dairesi = str(row[3]).strip() if len(row) > 3 and row[3] else None
                    
                    # Validasyon
                    if not vkn_tckn:
                        raise ValueError("VKN / TC Kimlik boş olamaz")
                    
                    if not validate_vkn_tckn(vkn_tckn):
                        raise ValueError(f"Geçersiz VKN/TC: {vkn_tckn} (10 veya 11 haneli rakam olmalı)")
                    
                    if not bayi_kodu:
                        raise ValueError("Bayi Kodu boş olamaz")
                    
                    if not bayi_adi:
                        raise ValueError("Bayi Adı boş olamaz")
                    
                    # Bayi kodu unique mi?
                    existing_bayi = db.query(Bayi).filter(Bayi.bayi_kodu == bayi_kodu).first()
                    if existing_bayi:
                        raise ValueError(f"Bayi kodu zaten mevcut: {bayi_kodu}")
                    
                    # VKN'ye göre user bul veya oluştur (Multi-Company: current_user'ın company_id'sini kullan)
                    user = get_or_create_user_by_vkn(db, vkn_tckn, current_user.company_id, bayi_adi, vergi_dairesi)
                    
                    # Yeni bayi oluştur
                    new_bayi = Bayi(
                        bayi_kodu=bayi_kodu,
                        vkn_tckn=vkn_tckn,
                        user_id=user.id,
                        bayi_adi=bayi_adi,
                        vergi_dairesi=vergi_dairesi,  # Vergi dairesi
                        bakiye=0.0,  # Varsayılan
                        donem=None   # Dönem bilgisi yok
                    )
                    
                    db.add(new_bayi)
                    batch_counter += 1
                    
                    # 10'ar 10'ar commit et
                    if batch_counter >= BATCH_SIZE:
                        db.commit()
                        batch_counter = 0
                    
                    basarili_count += 1
                    olusturulan_bayiler.append({
                        "satir": row_num,
                        "vkn_tckn": vkn_tckn,
                        "bayi_kodu": bayi_kodu,
                        "bayi_adi": bayi_adi
                    })
                    
                    # İlerleme gönder
                    current_index = row_num - 2  # 0-based index
                    progress_percent = int((current_index + 1) / total_rows * 100)
                    yield f"data: {json.dumps({'type': 'progress', 'current': current_index + 1, 'total': total_rows, 'percent': progress_percent, 'basarili': basarili_count, 'basarisiz': basarisiz_count})}\n\n"
                    
                except Exception as row_error:
                    basarisiz_count += 1
                    hatalar.append({
                        "satir": row_num,
                        "hata": str(row_error),
                        "veri": {
                            "vkn_tckn": row[0] if len(row) > 0 else None,
                            "bayi_kodu": row[1] if len(row) > 1 else None,
                            "bayi_adi": row[2] if len(row) > 2 else None
                        }
                    })
                    
                    # Hata gönder
                    current_index = row_num - 2
                    progress_percent = int((current_index + 1) / total_rows * 100)
                    yield f"data: {json.dumps({'type': 'progress', 'current': current_index + 1, 'total': total_rows, 'percent': progress_percent, 'basarili': basarili_count, 'basarisiz': basarisiz_count})}\n\n"
            
            # Son batch'i commit et
            if batch_counter > 0:
                db.commit()
            
            # Activity log
            if basarili_count > 0:
                ActivityLogger.log(
                    db, "BAYI_YUKLE_TOPLU", 
                    f"Toplu bayi yuklendi: {basarili_count} basarili, {basarisiz_count} basarisiz", 
                    current_user.id, 
                    "127.0.0.1"
                )
            
            # Tamamlandı mesajı
            yield f"data: {json.dumps({'type': 'complete', 'toplam': total_rows, 'basarili': basarili_count, 'basarisiz': basarisiz_count, 'hatalar': hatalar, 'olusturulan_bayiler': olusturulan_bayiler[:10]})}\n\n"
            
        except Exception as e:
            db.rollback()
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)[:200]})}\n\n"
    
    return StreamingResponse(
        generate_progress(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )

@router.get("/by-vkn/{vkn_tckn}", response_model=BayiSummaryByVKN)
def get_bayiler_by_vkn(
    vkn_tckn: str,
    donem: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    VKN'ye göre bayileri getir - Multi-Company (Sadece giriş yapılan şirketteki bayiler)
    """
    
    # VKN validasyonu
    if not validate_vkn_tckn(vkn_tckn):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Geçersiz VKN/TC Kimlik numarası"
        )
    
    # VKN + company_id ile kullanıcıyı bul (Multi-Company: Aynı VKN farklı şirketlerde olabilir)
    user = db.query(User).filter(
        User.vkn_tckn == vkn_tckn,
        User.company_id == current_user.company_id  # ÖNEMLİ: Şirket filtresi
    ).first()
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bu VKN'ye ait kayıt bulunamadı: {vkn_tckn} (Şirket: {current_user.company_id})"
        )
    
    # Sadece bu kullanıcıya ait bayileri getir
    query = db.query(Bayi).filter(Bayi.user_id == user.id)
    
    if donem:
        query = query.filter(Bayi.donem == donem)
    
    bayiler = query.all()
    
    if not bayiler:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bu VKN'ye ait bayi bulunamadı"
        )
    
    # Toplam bakiye hesapla
    toplam_bakiye = sum(b.bakiye for b in bayiler)
    
    return BayiSummaryByVKN(
        vkn_tckn=vkn_tckn,
        company_name=user.company_name,
        toplam_bayi_sayisi=len(bayiler),
        toplam_bakiye=toplam_bakiye,
        bayiler=[BayiResponse.from_orm(b) for b in bayiler]
    )

@router.get("/", response_model=List[BayiResponse])
def get_all_bayiler(
    skip: int = 0,
    limit: int = 100,
    vkn_tckn: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Tüm bayileri listele (pagination)
    """
    
    query = db.query(Bayi)
    
    if vkn_tckn:
        query = query.filter(Bayi.vkn_tckn == vkn_tckn)
    
    bayiler = query.offset(skip).limit(limit).all()
    
    return [BayiResponse.from_orm(b) for b in bayiler]

@router.delete("/{bayi_id}")
def delete_bayi(
    bayi_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Bayi kaydını sil (Multi-Company)
    """
    
    # Sadece admin ve company_admin silebilir
    if current_user.role not in [UserRole.ADMIN, UserRole.COMPANY_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Sadece admin silebilir"
        )
    
    bayi = db.query(Bayi).filter(Bayi.id == bayi_id).first()
    
    if not bayi:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Bayi bulunamadı"
        )
    
    db.delete(bayi)
    db.commit()
    
    ActivityLogger.log(
        db=db,
        action="BAYI_SIL",
        description=f"Bayi silindi: {bayi.bayi_kodu} - {bayi.bayi_adi}",
        user_id=current_user.id,
        ip_address="127.0.0.1"
    )
    
    return {"message": "Bayi başarıyla silindi"}


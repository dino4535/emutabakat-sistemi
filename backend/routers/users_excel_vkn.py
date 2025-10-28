"""
VKN Bazlı Toplu Kullanıcı ve Bayi Yükleme
Bir VKN'ye ait birden fazla bayi kodu tanımlanabilir
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Request
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
from backend.database import get_db
from backend.models import User, UserRole, Company, Bayi
from backend.auth import get_current_active_user
from backend.logger import ActivityLogger
from pydantic import BaseModel
import bcrypt
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

router = APIRouter(prefix="/api/auth", tags=["VKN Bazlı Excel"])


class VKNExcelUploadResult(BaseModel):
    toplam_satir: int
    basarili_user: int
    basarili_bayi: int
    basarisiz: int
    hatalar: List[dict]
    olusturulan_kullanicilar: List[dict]
    olusturulan_bayiler: List[dict]


def clean_phone(phone: str) -> str:
    """Telefon numarasını temizle"""
    if not phone:
        return ""
    phone = re.sub(r'[^\d]', '', str(phone))
    if phone and not phone.startswith('0'):
        phone = '0' + phone
    return phone[:11]


@router.post("/upload-users-vkn-excel", response_model=VKNExcelUploadResult)
def upload_users_vkn_excel(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    VKN Bazlı Toplu Kullanıcı ve Bayi Yükleme
    
    Excel Kolonları:
    1. VKN/TC (ZORUNLU)
    2. Bayi Kodu (ZORUNLU)
    3. Bayi Adı (ZORUNLU)
    4. Kullanıcı Adı (İlk VKN kaydı için ZORUNLU)
    5. E-posta (İlk VKN kaydı için ZORUNLU)
    6. Şifre (İlk VKN kaydı için ZORUNLU)
    7. Telefon
    8. Adres
    9. Rol (musteri/tedarikci)
    10. Şirket (Sistem admini için ZORUNLU)
    
    Mantık:
    - Her satır = 1 Bayi Kaydı
    - Aynı VKN için ilk satırda kullanıcı bilgileri dolu olmalı
    - Aynı VKN için sonraki satırlarda kullanıcı bilgileri boş bırakılabilir
    - Bir VKN'ye ait birden fazla bayi kodu (şube/mağaza) tanımlanabilir
    """
    
    # Yetki kontrolü
    if current_user.role not in [UserRole.ADMIN, UserRole.COMPANY_ADMIN]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Toplu kullanıcı yükleme yetkiniz yok"
        )
    
    # Dosya uzantısı kontrolü
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Lütfen Excel dosyası (.xlsx veya .xls) yükleyiniz"
        )
    
    try:
        # Excel dosyasını oku
        wb = load_workbook(file.file, data_only=True)
        ws = wb.active
        
        # İlk satır başlıklar
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel dosyası boş veya geçersiz format"
            )
        
        # Maksimum satır kontrolü
        MAX_ROWS = 1000
        if len(rows) > MAX_ROWS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Tek seferde maksimum {MAX_ROWS} satır yüklenebilir. Şu an {len(rows)} satır var."
            )
        
        basarili_user = 0
        basarili_bayi = 0
        basarisiz = 0
        hatalar = []
        olusturulan_kullanicilar = []
        olusturulan_bayiler = []
        
        # VKN cache (her VKN için bir kez user oluştur)
        vkn_user_cache = {}
        
        # Rol mapping
        rol_mapping = {
            "musteri": UserRole.MUSTERI,
            "müşteri": UserRole.MUSTERI,
            "tedarikci": UserRole.TEDARIKCI,
            "tedarikçi": UserRole.TEDARIKCI
        }
        
        for row_num, row in enumerate(rows, start=2):
            try:
                # Boş satırı atla
                if not any(row):
                    continue
                
                # Verileri al
                vkn_tckn = str(row[0]).strip() if row[0] else None
                bayi_kodu = str(row[1]).strip() if row[1] else None
                bayi_adi = str(row[2]).strip() if row[2] else None
                username = str(row[3]).strip() if row[3] else None
                email = str(row[4]).strip().lower() if row[4] else None
                password = str(row[5]).strip() if row[5] else None
                phone = clean_phone(row[6]) if row[6] else None
                address = str(row[7]).strip() if row[7] else None
                rol_str = str(row[8]).strip().lower() if row[8] else "musteri"
                company_identifier = str(row[9]).strip() if len(row) > 9 and row[9] else None
                
                # Zorunlu alan kontrolü
                if not vkn_tckn:
                    hatalar.append({
                        "satir": row_num,
                        "hata": "VKN/TC zorunludur",
                        "vkn": "-"
                    })
                    basarisiz += 1
                    continue
                
                if not bayi_kodu:
                    hatalar.append({
                        "satir": row_num,
                        "hata": "Bayi Kodu zorunludur",
                        "vkn": vkn_tckn
                    })
                    basarisiz += 1
                    continue
                
                if not bayi_adi:
                    hatalar.append({
                        "satir": row_num,
                        "hata": "Bayi Adı zorunludur",
                        "vkn": vkn_tckn
                    })
                    basarisiz += 1
                    continue
                
                # Rol kontrolü
                if rol_str not in rol_mapping:
                    hatalar.append({
                        "satir": row_num,
                        "hata": f"Geçersiz rol: {rol_str}. Olası değerler: musteri, tedarikci",
                        "vkn": vkn_tckn
                    })
                    basarisiz += 1
                    continue
                
                role = rol_mapping[rol_str]
                
                # Şirket ID belirleme
                target_company_id = None
                
                if current_user.role == UserRole.ADMIN:
                    # Sistem admini: Şirket kolonu zorunlu
                    if not company_identifier:
                        hatalar.append({
                            "satir": row_num,
                            "hata": "Sistem admini için 'Şirket' kolonu zorunludur",
                            "vkn": vkn_tckn
                        })
                        basarisiz += 1
                        continue
                    
                    # Şirket bul
                    target_company = db.query(Company).filter(
                        (Company.vkn == company_identifier) | (Company.company_name.ilike(f"%{company_identifier}%"))
                    ).first()
                    
                    if not target_company:
                        hatalar.append({
                            "satir": row_num,
                            "hata": f"Şirket bulunamadı: {company_identifier}",
                            "vkn": vkn_tckn
                        })
                        basarisiz += 1
                        continue
                    
                    target_company_id = target_company.id
                else:
                    # Şirket admini: Otomatik olarak kendi şirketi
                    target_company_id = current_user.company_id
                
                # VKN + Company ID ile cache key oluştur
                cache_key = f"{vkn_tckn}_{target_company_id}"
                
                # Bu VKN için user var mı kontrol et
                if cache_key not in vkn_user_cache:
                    # Database'de var mı?
                    existing_user = db.query(User).filter(
                        User.vkn_tckn == vkn_tckn,
                        User.company_id == target_company_id
                    ).first()
                    
                    if existing_user:
                        # Varolan kullanıcıyı cache'le
                        vkn_user_cache[cache_key] = existing_user
                    else:
                        # İlk kayıt: username, email, password gerekli
                        if not username or not email or not password:
                            hatalar.append({
                                "satir": row_num,
                                "hata": f"İlk VKN kaydı için Kullanıcı Adı, E-posta ve Şifre zorunludur (VKN: {vkn_tckn})",
                                "vkn": vkn_tckn
                            })
                            basarisiz += 1
                            continue
                        
                        # Email formatı kontrolü
                        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                        if not re.match(email_regex, email):
                            hatalar.append({
                                "satir": row_num,
                                "hata": f"Geçersiz e-posta formatı: {email}",
                                "vkn": vkn_tckn
                            })
                            basarisiz += 1
                            continue
                        
                        # Username ve email unique kontrolü
                        dup_user = db.query(User).filter(
                            (User.email == email) | (User.username == username)
                        ).first()
                        
                        if dup_user:
                            hatalar.append({
                                "satir": row_num,
                                "hata": f"Bu kullanıcı adı veya e-posta zaten kullanılıyor: {username} / {email}",
                                "vkn": vkn_tckn
                            })
                            basarisiz += 1
                            continue
                        
                        # Yeni kullanıcı oluştur
                        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                        
                        new_user = User(
                            company_id=target_company_id,
                            username=username,
                            email=email,
                            hashed_password=hashed_password,
                            full_name=bayi_adi,  # İlk bayi adını kullanıcı adı olarak kullan
                            company_name=bayi_adi,
                            vkn_tckn=vkn_tckn,
                            phone=phone,
                            address=address,
                            role=role,
                            is_active=True,
                            created_at=datetime.utcnow()
                        )
                        
                        db.add(new_user)
                        db.flush()
                        
                        vkn_user_cache[cache_key] = new_user
                        
                        olusturulan_kullanicilar.append({
                            "vkn": vkn_tckn,
                            "username": username,
                            "email": email,
                            "role": role.value
                        })
                        
                        basarili_user += 1
                
                # User referansı al
                user = vkn_user_cache[cache_key]
                
                # Bayi kaydı oluştur
                # Aynı bayi kodu var mı kontrol et
                existing_bayi = db.query(Bayi).filter(
                    Bayi.user_id == user.id,
                    Bayi.bayi_kodu == bayi_kodu
                ).first()
                
                if existing_bayi:
                    hatalar.append({
                        "satir": row_num,
                        "hata": f"Bu bayi kodu zaten kayıtlı: {bayi_kodu} (VKN: {vkn_tckn})",
                        "vkn": vkn_tckn
                    })
                    basarisiz += 1
                    continue
                
                new_bayi = Bayi(
                    user_id=user.id,
                    bayi_kodu=bayi_kodu,
                    bayi_adi=bayi_adi,
                    vkn_tckn=vkn_tckn,
                    bakiye=0.0,
                    donem=None,
                    created_at=datetime.utcnow()
                )
                
                db.add(new_bayi)
                db.flush()
                
                olusturulan_bayiler.append({
                    "vkn": vkn_tckn,
                    "bayi_kodu": bayi_kodu,
                    "bayi_adi": bayi_adi
                })
                
                basarili_bayi += 1
                
            except ValueError as e:
                hatalar.append({
                    "satir": row_num,
                    "hata": f"Geçersiz veri formatı: {str(e)}",
                    "vkn": vkn_tckn if vkn_tckn else "Bilinmiyor"
                })
                basarisiz += 1
            except Exception as e:
                hatalar.append({
                    "satir": row_num,
                    "hata": f"Beklenmeyen hata: {str(e)}",
                    "vkn": vkn_tckn if vkn_tckn else "Bilinmiyor"
                })
                basarisiz += 1
        
        # Tüm işlemler başarılı ise commit et
        db.commit()
        
        # Log
        ActivityLogger.log_activity(
            db,
            current_user.id,
            "TOPLU_VKN_BAYI_YUKLE",
            f"{basarili_user} kullanıcı, {basarili_bayi} bayi Excel'den yüklendi (VKN bazlı)",
            request.client.host if request.client else "unknown"
        )
        
        return VKNExcelUploadResult(
            toplam_satir=len(rows),
            basarili_user=basarili_user,
            basarili_bayi=basarili_bayi,
            basarisiz=basarisiz,
            hatalar=hatalar,
            olusturulan_kullanicilar=olusturulan_kullanicilar,
            olusturulan_bayiler=olusturulan_bayiler
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel işleme hatası: {str(e)}"
        )


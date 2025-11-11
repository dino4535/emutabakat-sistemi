"""Raporlama Router"""
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, and_, or_, text
from backend.database import get_db
from backend.models import User, Mutabakat, MutabakatDurumu, UserRole, ActivityLog
from backend.auth import get_current_active_user
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from fastapi.responses import StreamingResponse
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/api/reports", tags=["Reports"])

def require_admin(current_user: User = Depends(get_current_active_user)) -> User:
    """Admin yetkisi kontrolü - Multi-Company"""
    if current_user.role not in [UserRole.ADMIN, UserRole.COMPANY_ADMIN]:
        raise HTTPException(status_code=403, detail="Bu sayfaya erişim yetkiniz yok")
    return current_user

@router.get("/overview")
def get_overview_stats(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Genel İstatistikler - Multi-Company"""
    
    # Company ID bazlı filtreleme
    company_filter = User.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    # Toplam kullanıcılar
    total_users = db.query(User).filter(company_filter).count()
    active_users = db.query(User).filter(User.is_active == True, company_filter).count()
    
    # Kullanıcı rol dağılımı
    user_roles = db.query(
        User.role,
        func.count(User.id).label('count')
    ).filter(company_filter).group_by(User.role).all()
    
    # Toplam mutabakatlar
    total_mutabakats = db.query(Mutabakat).filter(mutabakat_company_filter).count()
    
    # Durum dağılımı
    status_distribution = db.query(
        Mutabakat.durum,
        func.count(Mutabakat.id).label('count')
    ).filter(mutabakat_company_filter).group_by(Mutabakat.durum).all()
    
    # Finansal özet
    financial_summary = db.query(
        func.sum(Mutabakat.toplam_borc).label('total_borc'),
        func.sum(Mutabakat.toplam_alacak).label('total_alacak'),
        func.sum(Mutabakat.bakiye).label('total_bakiye')
    ).filter(
        Mutabakat.durum == MutabakatDurumu.ONAYLANDI,
        mutabakat_company_filter
    ).first()
    
    # Son 30 günlük aktivite
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    recent_mutabakats = db.query(Mutabakat).filter(
        Mutabakat.created_at >= thirty_days_ago,
        mutabakat_company_filter
    ).count()
    
    return {
        "users": {
            "total": total_users,
            "active": active_users,
            "by_role": {role.value: count for role, count in user_roles}
        },
        "mutabakats": {
            "total": total_mutabakats,
            "recent_30_days": recent_mutabakats,
            "by_status": {status.value: count for status, count in status_distribution}
        },
        "financial": {
            "total_borc": float(financial_summary.total_borc or 0),
            "total_alacak": float(financial_summary.total_alacak or 0),
            "total_bakiye": float(financial_summary.total_bakiye or 0)
        }
    }

@router.get("/user-performance")
def get_user_performance(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Kullanıcı Performans Raporu - Multi-Company"""
    
    # Company ID bazlı filtreleme
    company_filter = User.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    # Her kullanıcının oluşturduğu mutabakat sayısı
    user_stats = db.query(
        User.id,
        User.username,
        User.full_name,
        User.company_name,
        User.role,
        func.count(Mutabakat.id).label('mutabakat_count')
    ).outerjoin(
        Mutabakat, and_(Mutabakat.sender_id == User.id, mutabakat_company_filter)
    ).filter(
        User.role.in_([UserRole.ADMIN, UserRole.COMPANY_ADMIN, UserRole.MUHASEBE, UserRole.PLANLAMA]),
        company_filter
    ).group_by(
        User.id, User.username, User.full_name, User.company_name, User.role
    ).all()
    
    result = []
    for stat in user_stats:
        # Her kullanıcının mutabakatlarının durum dağılımı
        status_dist = db.query(
            Mutabakat.durum,
            func.count(Mutabakat.id).label('count')
        ).filter(
            Mutabakat.sender_id == stat.id,
            mutabakat_company_filter
        ).group_by(Mutabakat.durum).all()
        
        result.append({
            "user_id": stat.id,
            "username": stat.username,
            "full_name": stat.full_name,
            "company_name": stat.company_name,
            "role": stat.role.value,
            "total_mutabakats": stat.mutabakat_count,
            "status_distribution": {status.value: count for status, count in status_dist}
        })
    
    return result

@router.get("/time-series")
def get_time_series(
    days: int = Query(90, ge=1, le=365, description="Kaç günlük zaman serisi"),
    group_by: str = Query("day", pattern="^(day|week)$", description="Gruplama: day|week"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Onay/Red zaman serisi ve ortalama onay süresi (günlük/haftalık)."""
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True

    start_dt = datetime.utcnow() - timedelta(days=days)

    # Veri çek
    q = db.query(Mutabakat).filter(
        Mutabakat.created_at >= start_dt,
        mutabakat_company_filter
    ).all()

    # Kovalara grupla
    from collections import defaultdict
    buckets = defaultdict(lambda: {"sent": 0, "approved": 0, "rejected": 0, "avg_response_days_sum": 0.0, "avg_response_days_cnt": 0})

    def bucket_key(dt: datetime) -> str:
        if group_by == "week":
            iso = dt.isocalendar()
            return f"{iso.year}-W{iso.week:02d}"
        return dt.strftime("%Y-%m-%d")

    for m in q:
        if m.gonderim_tarihi:
            k = bucket_key(m.gonderim_tarihi)
            buckets[k]["sent"] += 1
        if m.onay_tarihi:
            k = bucket_key(m.onay_tarihi)
            buckets[k]["approved"] += 1
            if m.gonderim_tarihi:
                diff_days = (m.onay_tarihi - m.gonderim_tarihi).days
                buckets[k]["avg_response_days_sum"] += max(diff_days, 0)
                buckets[k]["avg_response_days_cnt"] += 1
        if m.red_tarihi:
            k = bucket_key(m.red_tarihi)
            buckets[k]["rejected"] += 1

    # Sıralı liste
    def sort_key(x: str) -> tuple:
        if group_by == "week":
            year, w = x.split("-W")
            return (int(year), int(w))
        return tuple(map(int, x.split("-")))

    series = []
    for k in sorted(buckets.keys(), key=sort_key):
        b = buckets[k]
        avg_resp = (b["avg_response_days_sum"] / b["avg_response_days_cnt"]) if b["avg_response_days_cnt"] > 0 else 0.0
        series.append({
            "bucket": k,
            "sent": b["sent"],
            "approved": b["approved"],
            "rejected": b["rejected"],
            "avg_response_days": round(float(avg_resp), 2)
        })

    return {
        "group_by": group_by,
        "start": start_dt.isoformat(),
        "days": days,
        "series": series
    }

@router.get("/time-series/export.csv")
def export_time_series_csv(
    days: int = Query(90, ge=1, le=365),
    group_by: str = Query("day", pattern="^(day|week)$"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Zaman serisi CSV export."""
    data = get_time_series(days=days, group_by=group_by, current_user=current_user, db=db)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["bucket", "sent", "approved", "rejected", "avg_response_days"])
    for row in data["series"]:
        writer.writerow([row["bucket"], row["sent"], row["approved"], row["rejected"], row["avg_response_days"]])

    output.seek(0)
    filename = f"time_series_{group_by}_{days}d.csv"
    return StreamingResponse(iter([output.read()]), media_type="text/csv", headers={
        "Content-Disposition": f'attachment; filename="{filename}"'
    })

@router.get("/monthly-trend")
def get_monthly_trend(
    months: int = 6,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Aylık Trend Analizi - Multi-Company"""
    
    # Company ID bazlı filtreleme
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    # Son N ay için mutabakat sayıları
    monthly_data = db.query(
        extract('year', Mutabakat.created_at).label('year'),
        extract('month', Mutabakat.created_at).label('month'),
        func.count(Mutabakat.id).label('count'),
        func.sum(Mutabakat.toplam_borc).label('total_borc'),
        func.sum(Mutabakat.toplam_alacak).label('total_alacak')
    ).filter(mutabakat_company_filter).group_by(
        'year', 'month'
    ).order_by(
        'year', 'month'
    ).limit(months).all()
    
    result = []
    for data in monthly_data:
        result.append({
            "year": int(data.year),
            "month": int(data.month),
            "count": data.count,
            "total_borc": float(data.total_borc or 0),
            "total_alacak": float(data.total_alacak or 0)
        })
    
    return result

@router.get("/recent-activities")
def get_recent_activities(
    limit: int = 50,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Son Aktiviteler - Multi-Company"""
    
    # Company ID bazlı filtreleme - ActivityLog'da user_id üzerinden filtreleme
    if current_user.role == UserRole.COMPANY_ADMIN:
        # Sadece kendi şirketinin kullanıcılarının aktivitelerini göster
        company_user_ids = db.query(User.id).filter(User.company_id == current_user.company_id).all()
        company_user_ids = [uid[0] for uid in company_user_ids]
        activities = db.query(ActivityLog).filter(
            ActivityLog.user_id.in_(company_user_ids)
        ).order_by(
            ActivityLog.created_at.desc()
        ).limit(limit).all()
    else:
        activities = db.query(ActivityLog).order_by(
            ActivityLog.created_at.desc()
        ).limit(limit).all()
    
    result = []
    for activity in activities:
        user = db.query(User).filter(User.id == activity.user_id).first() if activity.user_id else None
        result.append({
            "id": activity.id,
            "action": activity.action,
            "description": activity.description,
            "username": user.username if user else "Sistem",
            "full_name": user.full_name if user else "Sistem",
            "ip_address": activity.ip_address,
            "created_at": activity.created_at.isoformat() if activity.created_at else None
        })
    
    return result

@router.get("/top-customers")
def get_top_customers(
    limit: int = 10,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """En Çok Mutabakat Alan Müşteriler - Multi-Company"""
    
    # Company ID bazlı filtreleme
    company_filter = User.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    top_customers = db.query(
        User.id,
        User.username,
        User.full_name,
        User.company_name,
        func.count(Mutabakat.id).label('mutabakat_count'),
        func.sum(Mutabakat.toplam_borc).label('total_borc'),
        func.sum(Mutabakat.toplam_alacak).label('total_alacak')
    ).join(
        Mutabakat, and_(Mutabakat.receiver_id == User.id, mutabakat_company_filter)
    ).filter(
        User.role.in_([UserRole.MUSTERI, UserRole.TEDARIKCI]),
        company_filter
    ).group_by(
        User.id, User.username, User.full_name, User.company_name
    ).order_by(
        func.count(Mutabakat.id).desc()
    ).limit(limit).all()
    
    result = []
    for customer in top_customers:
        # Onay oranı hesapla
        total = db.query(func.count(Mutabakat.id)).filter(
            Mutabakat.receiver_id == customer.id,
            mutabakat_company_filter
        ).scalar()
        
        approved = db.query(func.count(Mutabakat.id)).filter(
            Mutabakat.receiver_id == customer.id,
            Mutabakat.durum == MutabakatDurumu.ONAYLANDI,
            mutabakat_company_filter
        ).scalar()
        
        approval_rate = (approved / total * 100) if total > 0 else 0
        
        result.append({
            "user_id": customer.id,
            "username": customer.username,
            "full_name": customer.full_name,
            "company_name": customer.company_name,
            "mutabakat_count": customer.mutabakat_count,
            "total_borc": float(customer.total_borc or 0),
            "total_alacak": float(customer.total_alacak or 0),
            "approval_rate": round(approval_rate, 2)
        })
    
    return result

@router.get("/approval-statistics")
def get_approval_statistics(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Onay/Red İstatistikleri - Multi-Company"""
    
    # Company ID bazlı filtreleme
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    # Toplam gönderilmiş
    total_sent = db.query(Mutabakat).filter(
        Mutabakat.durum != MutabakatDurumu.TASLAK,
        mutabakat_company_filter
    ).count()
    
    # Onaylanan
    approved = db.query(Mutabakat).filter(
        Mutabakat.durum == MutabakatDurumu.ONAYLANDI,
        mutabakat_company_filter
    ).count()
    
    # Reddedilen
    rejected = db.query(Mutabakat).filter(
        Mutabakat.durum == MutabakatDurumu.REDDEDILDI,
        mutabakat_company_filter
    ).count()
    
    # Bekleyen
    pending = db.query(Mutabakat).filter(
        Mutabakat.durum == MutabakatDurumu.GONDERILDI,
        mutabakat_company_filter
    ).count()
    
    # Ortalama yanıt süresi (onaylananlar için)
    # SQL Server için DATEDIFF kullanıyoruz - Raw SQL ile
    if current_user.role == UserRole.COMPANY_ADMIN:
        avg_response_time_result = db.execute(
            text("""
                SELECT AVG(CAST(DATEDIFF(day, gonderim_tarihi, onay_tarihi) AS FLOAT)) 
                FROM mutabakats 
                WHERE durum = :durum 
                AND onay_tarihi IS NOT NULL 
                AND gonderim_tarihi IS NOT NULL
                AND company_id = :company_id
            """),
            {"durum": "ONAYLANDI", "company_id": current_user.company_id}
        ).scalar()
    else:
        avg_response_time_result = db.execute(
            text("""
                SELECT AVG(CAST(DATEDIFF(day, gonderim_tarihi, onay_tarihi) AS FLOAT)) 
                FROM mutabakats 
                WHERE durum = :durum 
                AND onay_tarihi IS NOT NULL 
                AND gonderim_tarihi IS NOT NULL
            """),
            {"durum": "ONAYLANDI"}
        ).scalar()
    
    avg_response_time = avg_response_time_result if avg_response_time_result else 0
    
    return {
        "total_sent": total_sent,
        "approved": approved,
        "rejected": rejected,
        "pending": pending,
        "approval_rate": round((approved / total_sent * 100) if total_sent > 0 else 0, 2),
        "rejection_rate": round((rejected / total_sent * 100) if total_sent > 0 else 0, 2),
        "avg_response_time_days": round(float(avg_response_time or 0), 2)
    }

@router.get("/detailed-user-analysis")
def get_detailed_user_analysis(
    user_id: Optional[int] = Query(None, description="Kullanıcı ID filtresi"),
    period: Optional[str] = Query(None, description="Dönem filtresi (MM/YYYY formatında, örn: 09/2025)"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Detaylı Kullanıcı Analizi (Dönem ve Kullanıcı Filtreli) - Multi-Company"""
    
    # Company ID bazlı filtreleme
    company_filter = User.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    # Dönem filtresini hazırla (donem_bitis alanına göre)
    date_filters = []
    if period:
        try:
            # MM/YYYY formatından ayrıştır
            month_str, year_str = period.split('/')
            month = int(month_str)
            year = int(year_str)
            
            # O ayın başlangıç ve bitiş tarihleri
            start_of_month = datetime(year, month, 1, 0, 0, 0)
            if month == 12:
                end_of_month = datetime(year + 1, 1, 1, 0, 0, 0)
            else:
                end_of_month = datetime(year, month + 1, 1, 0, 0, 0)
            
            # donem_bitis alanı bu ay içinde olan mutabakatlar
            date_filters.append(Mutabakat.donem_bitis >= start_of_month)
            date_filters.append(Mutabakat.donem_bitis < end_of_month)
        except (ValueError, AttributeError):
            pass
    
    # Kullanıcı filtresi
    user_filter = []
    if user_id:
        user_filter.append(User.id == user_id)
    
    # Şirket içi kullanıcıları al (admin, company_admin, muhasebe, planlama)
    users_query = db.query(User).filter(
        User.role.in_([UserRole.ADMIN, UserRole.COMPANY_ADMIN, UserRole.MUHASEBE, UserRole.PLANLAMA]),
        company_filter
    )
    
    if user_filter:
        users_query = users_query.filter(*user_filter)
    
    users = users_query.all()
    
    result = []
    for user in users:
        # Temel sorgu
        base_query = db.query(Mutabakat).filter(
            Mutabakat.sender_id == user.id,
            mutabakat_company_filter
        )
        
        # Tarih filtrelerini uygula
        if date_filters:
            base_query = base_query.filter(and_(*date_filters))
        
        # Toplam oluşturulan
        total_created = base_query.count()
        
        # Gönderilen (Taslak olmayan)
        total_sent = base_query.filter(Mutabakat.durum != MutabakatDurumu.TASLAK).count()
        
        # Cevaplanmış (Onaylanmış veya Reddedilmiş)
        total_answered = base_query.filter(
            or_(
                Mutabakat.durum == MutabakatDurumu.ONAYLANDI,
                Mutabakat.durum == MutabakatDurumu.REDDEDILDI
            )
        ).count()
        
        # Onaylanan
        total_approved = base_query.filter(Mutabakat.durum == MutabakatDurumu.ONAYLANDI).count()
        
        # Reddedilen
        total_rejected = base_query.filter(Mutabakat.durum == MutabakatDurumu.REDDEDILDI).count()
        
        # Bekleyen
        total_pending = base_query.filter(Mutabakat.durum == MutabakatDurumu.GONDERILDI).count()
        
        # Taslak
        total_draft = base_query.filter(Mutabakat.durum == MutabakatDurumu.TASLAK).count()
        
        # Finansal toplamlar (onaylanmış mutabakatlar için)
        financial = base_query.filter(Mutabakat.durum == MutabakatDurumu.ONAYLANDI).all()
        total_borc = sum(m.toplam_borc or 0 for m in financial)
        total_alacak = sum(m.toplam_alacak or 0 for m in financial)
        total_bakiye = sum(m.bakiye or 0 for m in financial)
        
        # Ortalama yanıt süresi (sadece bu kullanıcının onaylanmış mutabakatları için)
        avg_response = base_query.filter(
            Mutabakat.durum == MutabakatDurumu.ONAYLANDI,
            Mutabakat.onay_tarihi.isnot(None),
            Mutabakat.gonderim_tarihi.isnot(None)
        ).all()
        
        if avg_response:
            response_times = []
            for m in avg_response:
                if m.onay_tarihi and m.gonderim_tarihi:
                    diff = (m.onay_tarihi - m.gonderim_tarihi).days
                    response_times.append(diff)
            avg_response_days = sum(response_times) / len(response_times) if response_times else 0
        else:
            avg_response_days = 0
        
        # Onay oranı
        approval_rate = (total_approved / total_sent * 100) if total_sent > 0 else 0
        
        # Cevap oranı
        answer_rate = (total_answered / total_sent * 100) if total_sent > 0 else 0
        
        result.append({
            "user_id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "company_name": user.company_name,
            "role": user.role.value,
            "statistics": {
                "total_created": total_created,
                "total_sent": total_sent,
                "total_answered": total_answered,
                "total_approved": total_approved,
                "total_rejected": total_rejected,
                "total_pending": total_pending,
                "total_draft": total_draft,
                "approval_rate": round(approval_rate, 2),
                "answer_rate": round(answer_rate, 2),
                "avg_response_days": round(avg_response_days, 2)
            },
            "financial": {
                "total_borc": float(total_borc),
                "total_alacak": float(total_alacak),
                "total_bakiye": float(total_bakiye)
            }
        })
    
    # Toplam istatistikler
    total_stats = {
        "total_created": sum(u["statistics"]["total_created"] for u in result),
        "total_sent": sum(u["statistics"]["total_sent"] for u in result),
        "total_answered": sum(u["statistics"]["total_answered"] for u in result),
        "total_approved": sum(u["statistics"]["total_approved"] for u in result),
        "total_rejected": sum(u["statistics"]["total_rejected"] for u in result),
        "total_pending": sum(u["statistics"]["total_pending"] for u in result),
        "total_draft": sum(u["statistics"]["total_draft"] for u in result),
    }
    
    total_financial = {
        "total_borc": sum(u["financial"]["total_borc"] for u in result),
        "total_alacak": sum(u["financial"]["total_alacak"] for u in result),
        "total_bakiye": sum(u["financial"]["total_bakiye"] for u in result)
    }
    
    return {
        "users": result,
        "totals": {
            "statistics": total_stats,
            "financial": total_financial
        },
        "filters": {
            "user_id": user_id,
            "period": period
        }
    }

@router.get("/available-periods")
def get_available_periods(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Sistemdeki mevcut dönemleri listele (donem_bitis'e göre) - Multi-Company"""
    
    # Company ID bazlı filtreleme
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    # Tüm mutabakatların donem_bitis tarihlerini al
    mutabakats = db.query(Mutabakat).filter(
        Mutabakat.donem_bitis.isnot(None),
        mutabakat_company_filter
    ).all()
    
    # Benzersiz ay/yıl kombinasyonları
    periods = set()
    for m in mutabakats:
        if m.donem_bitis:
            period = f"{m.donem_bitis.month:02d}/{m.donem_bitis.year}"
            periods.add(period)
    
    # Sırala (yeniden eskiye)
    sorted_periods = sorted(list(periods), key=lambda x: (int(x.split('/')[1]), int(x.split('/')[0])), reverse=True)
    
    return sorted_periods

@router.get("/period-comparison")
def get_period_comparison(
    end_period: Optional[str] = Query(None, description="Bitiş dönemi MM/YYYY (ör. 12/2025). Belirtilmezse bugüne göre hesaplanır."),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Dönemsel Karşılaştırma - Son 12 Ay (Dönem Bitiş Tarihine Göre) - Multi-Company"""
    
    # Company ID bazlı filtreleme
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    result = []

    # Referans ay belirle
    if end_period:
        try:
            month_str, year_str = end_period.split('/')
            ref_month = int(month_str)
            ref_year = int(year_str)
            reference = datetime(ref_year, ref_month, 1, 0, 0, 0)
        except Exception:
            reference = datetime.utcnow()
    else:
        reference = datetime.utcnow()

    for i in range(0, 12):  # Son 12 ay (referans ay -> 11 ay önce)
        # Dönem başlangıç ve bitiş tarihleri
        target_date = reference - timedelta(days=30*i)
        start_of_month = target_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Bir sonraki ayın ilk günü
        if start_of_month.month == 12:
            end_of_month = start_of_month.replace(year=start_of_month.year + 1, month=1)
        else:
            end_of_month = start_of_month.replace(month=start_of_month.month + 1)
        
        # O ay dönemine ait mutabakatlar (dönem bitiş tarihine göre)
        monthly_mutabakats = db.query(Mutabakat).filter(
            and_(
                Mutabakat.donem_bitis.isnot(None),  # Dönem bitişi olan mutabakatlar
                Mutabakat.donem_bitis >= start_of_month,
                Mutabakat.donem_bitis < end_of_month,
                mutabakat_company_filter
            )
        ).all()
        
        # İstatistikler
        total = len(monthly_mutabakats)
        approved = len([m for m in monthly_mutabakats if m.durum == MutabakatDurumu.ONAYLANDI])
        rejected = len([m for m in monthly_mutabakats if m.durum == MutabakatDurumu.REDDEDILDI])
        pending = len([m for m in monthly_mutabakats if m.durum == MutabakatDurumu.GONDERILDI])
        draft = len([m for m in monthly_mutabakats if m.durum == MutabakatDurumu.TASLAK])
        
        # Finansal toplamlar
        approved_mutabakats = [m for m in monthly_mutabakats if m.durum == MutabakatDurumu.ONAYLANDI]
        total_borc = sum(m.toplam_borc or 0 for m in approved_mutabakats)
        total_alacak = sum(m.toplam_alacak or 0 for m in approved_mutabakats)
        
        result.append({
            "year": start_of_month.year,
            "month": start_of_month.month,
            "month_name": start_of_month.strftime("%B %Y"),
            "total": total,
            "approved": approved,
            "rejected": rejected,
            "pending": pending,
            "draft": draft,
            "total_borc": float(total_borc),
            "total_alacak": float(total_alacak),
            "approval_rate": round((approved / total * 100) if total > 0 else 0, 2)
        })
    
    return result

@router.get("/pending-heatmap")
def get_pending_heatmap(
    days: int = Query(30, description="Son kaç günlük veri (default: 30)"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Bekleyen Mutabakat Isı Haritası - Multi-Company
    
    Returns:
    - Günlük dağılım (hangi günlerde daha fazla bekleyen var)
    - Bekleme süresine göre gruplandırma
    - Müşteri bazlı bekleme süreleri
    - Renk kodlaması için veriler
    """
    from collections import defaultdict
    
    # Company ID bazlı filtreleme
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    # Son N gün içinde gönderilen ama henüz onaylanmayan mutabakatlar
    date_threshold = datetime.utcnow() - timedelta(days=days)
    
    pending_mutabakats = db.query(Mutabakat).filter(
        Mutabakat.durum == MutabakatDurumu.GONDERILDI,
        Mutabakat.gonderim_tarihi >= date_threshold,
        mutabakat_company_filter
    ).all()
    
    # Günlük dağılım (gönderim tarihine göre)
    daily_distribution = defaultdict(int)
    daily_waiting_times = defaultdict(list)
    
    # Bekleme süresine göre gruplandırma
    waiting_buckets = {
        "0-1": [],  # 0-1 gün
        "1-3": [],  # 1-3 gün
        "3-7": [],  # 3-7 gün
        "7-14": [], # 7-14 gün
        "14+": []   # 14+ gün
    }
    
    # Müşteri bazlı bekleme süreleri
    customer_waiting = defaultdict(lambda: {"count": 0, "total_days": 0, "max_days": 0})
    
    now = datetime.utcnow()
    
    for mutabakat in pending_mutabakats:
        if not mutabakat.gonderim_tarihi:
            continue
        
        # Bekleme süresi (gün cinsinden)
        waiting_days = (now - mutabakat.gonderim_tarihi).days
        
        # Günlük dağılım
        send_date = mutabakat.gonderim_tarihi.date()
        daily_distribution[str(send_date)] += 1
        daily_waiting_times[str(send_date)].append(waiting_days)
        
        # Bekleme süresine göre gruplandırma
        if waiting_days <= 1:
            waiting_buckets["0-1"].append(mutabakat.id)
        elif waiting_days <= 3:
            waiting_buckets["1-3"].append(mutabakat.id)
        elif waiting_days <= 7:
            waiting_buckets["3-7"].append(mutabakat.id)
        elif waiting_days <= 14:
            waiting_buckets["7-14"].append(mutabakat.id)
        else:
            waiting_buckets["14+"].append(mutabakat.id)
        
        # Müşteri bazlı
        if mutabakat.receiver:
            receiver_name = mutabakat.receiver.company_name or mutabakat.receiver.full_name or mutabakat.receiver.username
            customer_waiting[receiver_name]["count"] += 1
            customer_waiting[receiver_name]["total_days"] += waiting_days
            customer_waiting[receiver_name]["max_days"] = max(customer_waiting[receiver_name]["max_days"], waiting_days)
    
    # Günlük dağılımı formatla (renk kodlaması ile)
    daily_data = []
    for date_str, count in sorted(daily_distribution.items()):
        avg_waiting = sum(daily_waiting_times[date_str]) / len(daily_waiting_times[date_str]) if daily_waiting_times[date_str] else 0
        
        # Renk kodlaması: kırmızı = uzun süre, yeşil = yeni
        if avg_waiting <= 1:
            color = "green"
            intensity = 1
        elif avg_waiting <= 3:
            color = "yellow"
            intensity = 2
        elif avg_waiting <= 7:
            color = "orange"
            intensity = 3
        else:
            color = "red"
            intensity = 4
        
        daily_data.append({
            "date": date_str,
            "count": count,
            "avg_waiting_days": round(avg_waiting, 1),
            "color": color,
            "intensity": intensity
        })
    
    # Müşteri bazlı verileri formatla
    customer_data = []
    for customer_name, data in customer_waiting.items():
        avg_days = data["total_days"] / data["count"] if data["count"] > 0 else 0
        
        # Renk kodlaması
        if avg_days <= 1:
            color = "green"
        elif avg_days <= 3:
            color = "yellow"
        elif avg_days <= 7:
            color = "orange"
        else:
            color = "red"
        
        customer_data.append({
            "customer_name": customer_name,
            "pending_count": data["count"],
            "avg_waiting_days": round(avg_days, 1),
            "max_waiting_days": data["max_days"],
            "color": color
        })
    
    # En uzun bekleyenler (top 10)
    longest_waiting = sorted(
        [
            {
                "mutabakat_no": m.mutabakat_no,
                "receiver_name": m.receiver.company_name if m.receiver else "Bilinmiyor",
                "waiting_days": (now - m.gonderim_tarihi).days,
                "send_date": m.gonderim_tarihi.isoformat() if m.gonderim_tarihi else None,
                "amount": float(m.bakiye) if m.bakiye else 0.0
            }
            for m in pending_mutabakats
            if m.gonderim_tarihi
        ],
        key=lambda x: x["waiting_days"],
        reverse=True
    )[:10]
    
    return {
        "total_pending": len(pending_mutabakats),
        "waiting_buckets": {
            bucket: len(ids) for bucket, ids in waiting_buckets.items()
        },
        "daily_distribution": daily_data,
        "customer_waiting": sorted(customer_data, key=lambda x: x["avg_waiting_days"], reverse=True),
        "longest_waiting": longest_waiting,
        "summary": {
            "avg_waiting_days": round(
                sum((now - m.gonderim_tarihi).days for m in pending_mutabakats if m.gonderim_tarihi) / len([m for m in pending_mutabakats if m.gonderim_tarihi]),
                1
            ) if any(m.gonderim_tarihi for m in pending_mutabakats) else 0,
            "max_waiting_days": max(
                ((now - m.gonderim_tarihi).days for m in pending_mutabakats if m.gonderim_tarihi),
                default=0
            )
        }
    }

@router.get("/day-hour-heatmap")
def get_day_hour_heatmap(
    days: int = Query(30, ge=1, le=365),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Gün-saat ısı haritası (7x24) - gönderilmiş ama henüz cevaplanmamış mutabakatların dağılımı.
    day: 0=Monday ... 6=Sunday
    hour: 0..23
    """
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    start_dt = datetime.utcnow() - timedelta(days=days)

    q = db.query(Mutabakat).filter(
        Mutabakat.gonderim_tarihi.isnot(None),
        Mutabakat.gonderim_tarihi >= start_dt,
        Mutabakat.durum == MutabakatDurumu.GONDERILDI,
        mutabakat_company_filter
    ).all()

    # 7x24 matris
    matrix = [[0 for _ in range(24)] for _ in range(7)]

    for m in q:
        dt = m.gonderim_tarihi
        # Python: Monday=0 .. Sunday=6
        d = dt.weekday()
        h = dt.hour
        matrix[d][h] += 1

    # Normalize ve toplamları hesapla
    max_val = max((v for row in matrix for v in row), default=0)
    totals_by_day = [sum(row) for row in matrix]
    totals_by_hour = [sum(matrix[d][h] for d in range(7)) for h in range(24)]

    return {
        "days": days,
        "max": max_val,
        "matrix": matrix,
        "totals_by_day": totals_by_day,
        "totals_by_hour": totals_by_hour
    }

@router.get("/mutabakat-list/export.xlsx")
def export_mutabakat_list_excel(
    period: Optional[str] = Query(None, description="Dönem filtresi (MM/YYYY formatında, örn: 11/2025)"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Mutabakat listesini Excel olarak export et (dönem filtresi ile)
    Alıcının VKN'si dahil
    """
    # Company ID bazlı filtreleme
    mutabakat_company_filter = Mutabakat.company_id == current_user.company_id if current_user.role == UserRole.COMPANY_ADMIN else True
    
    # Dönem filtresini hazırla (donem_bitis alanına göre)
    date_filters = []
    if period:
        try:
            # MM/YYYY formatından ayrıştır
            month_str, year_str = period.split('/')
            month = int(month_str)
            year = int(year_str)
            
            # O ayın başlangıç ve bitiş tarihleri
            start_of_month = datetime(year, month, 1, 0, 0, 0)
            if month == 12:
                end_of_month = datetime(year + 1, 1, 1, 0, 0, 0)
            else:
                end_of_month = datetime(year, month + 1, 1, 0, 0, 0)
            
            # donem_bitis alanı bu ay içinde olan mutabakatlar
            date_filters.append(Mutabakat.donem_bitis >= start_of_month)
            date_filters.append(Mutabakat.donem_bitis < end_of_month)
        except (ValueError, AttributeError):
            pass
    
    # Mutabakatları çek
    query = db.query(Mutabakat).filter(
        mutabakat_company_filter
    )
    
    if date_filters:
        query = query.filter(and_(*date_filters))
    
    mutabakats = query.order_by(Mutabakat.created_at.desc()).all()
    
    # Excel workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Mutabakat Listesi"
    
    # Stil tanımlamaları
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = [
        "Mutabakat No",
        "Oluşturan",
        "Gönderen/Alıcı",
        "Alıcının VKN'si",
        "Dönem Başlangıç",
        "Dönem Bitiş",
        "Durum",
        "Borç (₺)",
        "Alacak (₺)",
        "Bakiye (₺)",
        "Toplam Bayi Sayısı",
        "Gönderim Tarihi",
        "Onay Tarihi",
        "Red Tarihi",
        "Oluşturulma Tarihi"
    ]
    
    # Başlıkları yaz
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Verileri yaz
    for row_idx, mutabakat in enumerate(mutabakats, start=2):
        # Durum metni
        durum_text = {
            MutabakatDurumu.TASLAK: "Taslak",
            MutabakatDurumu.GONDERILDI: "Gönderildi",
            MutabakatDurumu.ONAYLANDI: "Onaylandı",
            MutabakatDurumu.REDDEDILDI: "Reddedildi",
            MutabakatDurumu.IPTAL: "İptal"
        }.get(mutabakat.durum, str(mutabakat.durum))
        
        # Kullanıcı bilgileri
        sender_name = mutabakat.sender.company_name or mutabakat.sender.full_name or mutabakat.sender.username if mutabakat.sender else "-"
        receiver_name = mutabakat.receiver.company_name or mutabakat.receiver.full_name or mutabakat.receiver.username if mutabakat.receiver else "-"
        receiver_vkn = mutabakat.receiver_vkn if mutabakat.receiver_vkn else (mutabakat.receiver.vkn_tckn if mutabakat.receiver else "-")
        
        # Tarih formatlama
        def format_date(dt):
            if dt:
                return dt.strftime("%d.%m.%Y %H:%M")
            return "-"
        
        data = [
            mutabakat.mutabakat_no,
            sender_name,
            receiver_name,
            receiver_vkn,
            format_date(mutabakat.donem_baslangic),
            format_date(mutabakat.donem_bitis),
            durum_text,
            round(mutabakat.toplam_borc or 0, 2),
            round(mutabakat.toplam_alacak or 0, 2),
            round(mutabakat.bakiye or 0, 2),
            mutabakat.toplam_bayi_sayisi or 0,
            format_date(mutabakat.gonderim_tarihi),
            format_date(mutabakat.onay_tarihi),
            format_date(mutabakat.red_tarihi),
            format_date(mutabakat.created_at)
        ]
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Sayısal sütunlar için sağa hizalama
            if col_idx in [8, 9, 10, 11]:  # Borç, Alacak, Bakiye, Bayi Sayısı
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_idx in [8, 9, 10]:  # Para birimi sütunları
                    cell.number_format = '#,##0.00'
    
    # Sütun genişliklerini ayarla
    column_widths = {
        'A': 20,  # Mutabakat No
        'B': 30,  # Oluşturan
        'C': 30,  # Gönderen/Alıcı
        'D': 15,  # Alıcının VKN'si
        'E': 18,  # Dönem Başlangıç
        'F': 18,  # Dönem Bitiş
        'G': 12,  # Durum
        'H': 15,  # Borç
        'I': 15,  # Alacak
        'J': 15,  # Bakiye
        'K': 15,  # Bayi Sayısı
        'L': 18,  # Gönderim Tarihi
        'M': 18,  # Onay Tarihi
        'N': 18,  # Red Tarihi
        'O': 18   # Oluşturulma Tarihi
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Excel dosyasını memory'de oluştur
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Dosya adı
    period_str = period.replace('/', '-') if period else "tum-donemler"
    filename = f"mutabakat-listesi-{period_str}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


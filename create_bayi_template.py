"""
Toplu Bayi Yükleme Template Oluşturma Scripti
Sadece 3 kolon: VKN/TC, Bayi Kodu, Bayi Adı
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_template():
    wb = openpyxl.Workbook()
    ws_list = wb.active
    ws_list.title = "Bayi Listesi"

    # 4 kolon: VKN/TC, Bayi Kodu, Bayi Adı, Vergi Dairesi
    headers = ["VKN/TC", "Bayi Kodu", "Bayi Adı", "Vergi Dairesi"]
    ws_list.append(headers)

    # Header stilleri
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_list.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Kolon genişlikleri
        if col_num == 1:  # VKN/TC
            ws_list.column_dimensions[get_column_letter(col_num)].width = 15
        elif col_num == 2:  # Bayi Kodu
            ws_list.column_dimensions[get_column_letter(col_num)].width = 15
        elif col_num == 3:  # Bayi Adı
            ws_list.column_dimensions[get_column_letter(col_num)].width = 40
        else:  # Vergi Dairesi
            ws_list.column_dimensions[get_column_letter(col_num)].width = 25

    # Örnek veriler
    example_data = [
        ["1234567890", "BY001", "ABC Marketleri Ankara Şubesi", "Çankaya Vergi Dairesi"],
        ["1234567890", "BY002", "ABC Marketleri İzmir Şubesi", "Konak Vergi Dairesi"],
        ["9876543210", "BY003", "XYZ Gıda A.Ş.", "Kadıköy Vergi Dairesi"],
        ["1122334455", "BY004", "Deneme Ticaret Ltd. Şti.", "Beşiktaş Vergi Dairesi"],
        ["66778899001", "BY005", "Örnek Sanayi ve Ticaret A.Ş.", "Ümraniye Vergi Dairesi"],
    ]
    
    for row_data in example_data:
        ws_list.append(row_data)
    
    # Kullanım kılavuzu sayfası
    ws_guide = wb.create_sheet("Kullanim Kilavuzu")
    guide_text = [
        "TOPLU BAYI YUKLEME KILAVUZU",
        "",
        "Bu Excel dosyasi, sisteme toplu bayi (cari kart) yuklemek icin kullanilir.",
        "Lutfen asagidaki kurallara dikkat ediniz:",
        "",
        "KOLONLAR:",
        "========",
        "1. VKN/TC: Vergi Kimlik Numarasi (10 hane) veya TC Kimlik Numarasi (11 hane) [ZORUNLU]",
        "   - Ayni VKN/TC'ye sahip birden fazla bayi olabilir.",
        "   - Eger VKN/TC sisteme kayitli degilse, otomatik kullanici hesabi olusturulur.",
        "   - Otomatik olusan kullanicilarin sifresi: VKN/TC'nin son 6 hanesi",
        "",
        "2. Bayi Kodu: Her bayi icin benzersiz bir kod [ZORUNLU]",
        "   - Ornekler: BY001, ANKA-001, MUS-123",
        "",
        "3. Bayi Adi: Bayinin tam adi veya unvani [ZORUNLU]",
        "   - Ornekler: ABC Marketleri Ankara Subesi, XYZ Gida A.S.",
        "",
        "4. Vergi Dairesi: Bayinin bagli oldugu vergi dairesi [OPSIYONEL]",
        "   - Ornekler: Cankaya Vergi Dairesi, Kadikoy Vergi Dairesi",
        "",
        "KURALLAR:",
        "=========",
        "- Maksimum 5000 satir yuklenebilir.",
        "- Dosya boyutu maksimum 5 MB olmalidir.",
        "- Kolon basliklari degistirilmemelidir.",
        "- Her satir icin tum 3 kolon doldurulmalidir.",
        "",
        "YUKLEME SONRASI:",
        "================",
        "- Basarili ve basarisiz kayitlar raporlanacaktir.",
        "- Hatali satirlar icin detayli aciklama gosterilecektir.",
        "- Yuklenen bayiler, VKN bazli mutabakat olusturma ekraninda listelenecektir.",
        "",
        "ONEMLI:",
        "=======",
        "- Ayni bayi kodunu birden fazla kez yukleyemezsiniz.",
        "- VKN/TC numara formatina dikkat ediniz (10 veya 11 hane).",
        "- Yeni olusturulan kullanicilar ilk giriste telefon ve email tamamlamalidir.",
    ]
    
    for i, line in enumerate(guide_text, 1):
        cell = ws_guide.cell(row=i, column=1)
        cell.value = line
        
        if i == 1:  # Baslik
            cell.font = Font(name='Arial', size=14, bold=True)
        elif line.endswith(":") and len(line) < 20:  # Alt basliklar
            cell.font = Font(name='Arial', size=11, bold=True)
        else:
            cell.font = Font(name='Arial', size=10)
        
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws_guide.column_dimensions['A'].width = 100

    filename = "template_bayiler.xlsx"
    wb.save(filename)
    print(f"[OK] Template olusturuldu: {filename}")
    print(f"     Kolonlar: VKN/TC | Bayi Kodu | Bayi Adi | Vergi Dairesi")
    print(f"     Ornek satir sayisi: 5")
    return filename

if __name__ == "__main__":
    create_template()
